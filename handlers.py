import logging
import os
import re
import json
import sqlite3
import shutil
from datetime import datetime
from pathlib import Path
from typing import List, Tuple, Sequence, Dict

from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.constants import ParseMode
from telegram.error import BadRequest
from telegram.ext import ContextTypes

import exhibitledger as el
from sheets_integration import format_sheets_status_markdown, format_sync_preview_markdown

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Utility Helpers (Self-contained to avoid circular imports with main.py)
# ---------------------------------------------------------------------------

def _parse_float(raw: str, label: str) -> float:
    try:
        return float(str(raw).replace(",", ""))
    except ValueError:
        raise ValueError(f"{label} must be a number.")

def _looks_number(token: str) -> bool:
    try:
        float(str(token).replace(",", ""))
        return True
    except ValueError:
        return False

def _pipe_parts(text: str) -> List[str]:
    return [part.strip() for part in (text or "").split("|")]

def _account_name_from_text(raw: str) -> str:
    names = el.account_head_names()
    cleaned = (raw or "").strip()
    if cleaned.isdigit():
        index = int(cleaned) - 1
        if 0 <= index < len(names):
            return names[index]
    lowered = cleaned.lower()
    for name in names:
        if lowered == name.lower():
            return name
    partial = [name for name in names if lowered and lowered in name.lower()]
    if len(partial) == 1:
        return partial[0]
    raise ValueError("Unknown account head. Use /accounts or choose from the menu.")

def _extract_receipt_code_and_text(chat_id: int, text: str) -> Tuple[str, str]:
    parts = (text or "").strip().split(maxsplit=1)
    if parts and el.get_exhibition(parts[0]):
        code = el.normalize_code(parts[0])
        rest = parts[1] if len(parts) > 1 else ""
        el.set_user_exhibition(chat_id, code)
        return code, rest
    state = el.get_user_state(chat_id)
    current = state.get("current_exhibition") or el.resolve_default_exhibition()
    return current, text

def _parse_split_tokens(code: str, tokens: List[str]) -> Tuple[str, List[dict]]:
    ALLOWED_SPLIT_TYPES = {"gallery", "artist", "collaborator", "collector"}
    if len(tokens) < 2:
        raise ValueError("Use: gallery 50 artist 50, or gallery 45 collaborator Curator 10 artist 45")
    entries = []
    idx = 0
    while idx < len(tokens):
        party_type = tokens[idx].lower().strip()
        if party_type not in ALLOWED_SPLIT_TYPES:
            raise ValueError(f"Unknown party type: {tokens[idx]}. Use gallery, artist, collaborator, or collector.")
        idx += 1
        if idx >= len(tokens):
            raise ValueError(f"Missing percentage after {party_type}.")
        party_name = party_type.title()
        if _looks_number(tokens[idx]):
            percent = _parse_float(tokens[idx], "Split percent")
            idx += 1
        else:
            party_name = tokens[idx]
            idx += 1
            if idx >= len(tokens):
                raise ValueError(f"Missing percentage after {party_type} {party_name}.")
            percent = _parse_float(tokens[idx], "Split percent")
            idx += 1
        entries.append({"party_type": party_type, "party_name": party_name, "percent": percent})
    return el.normalize_code(code), entries

def _parse_split_args(args: List[str]) -> Tuple[str, List[dict]]:
    if len(args) < 3:
        raise ValueError(
            "Usage: /set_split <EXHIBITION_CODE> gallery 50 artist 50\n"
            "Collaborator example: /set_split EXH gallery 40 collaborator Rotary 10 artist 50"
        )
    return _parse_split_tokens(el.normalize_code(args[0]), args[1:])

def generate_artwork_template_xlsx() -> str:
    """Generate and return path to the artwork bulk import template Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Artworks"

    headers = ["Title", "Artist", "Asking Price THB", "Notes (optional)"]
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    samples = [
        ["Sample Title 1 (delete this row)", "Sample Artist (delete this row)", 100000, ""],
        ["Sample Title 2 (delete this row)", "Sample Artist (delete this row)", 250000, "Replace these rows with your real artworks"],
    ]
    for row in samples:
        ws.append(row)

    ws2 = wb.create_sheet("Instructions")
    ws2.append(["Column", "Required", "Notes"])
    ws2.append(["Title", "Yes", "Artwork title"])
    ws2.append(["Artist", "Yes", "Artist full name"])
    ws2.append(["Asking Price THB", "Yes", "Number only, no commas or currency symbols"])
    ws2.append(["Notes (optional)", "No", "Any extra notes about the artwork"])
    ws2.append(["", "", ""])
    ws2.append(["", "", "Delete the 2 sample rows before uploading."])
    ws2.append(["", "", "Upload this file back to the bot after filling it in."])

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 35

    export_dir = os.environ.get("EXPORT_DIR", "./exports")
    os.makedirs(export_dir, exist_ok=True)
    path = os.path.join(export_dir, "artwork_import_template.xlsx")
    wb.save(path)
    return path

# Helper to log actions safely
def _log_action_safely(action: str, code: str | None, details: str) -> None:
    try:
        el.log_action(action, code, details)
    except Exception as e:
        logger.warning(f"Failed to write audit log: {e}")

# Context / state resolution
def _get_current_exhibition(chat_id: int) -> str:
    state = el.get_user_state(chat_id)
    return state.get("current_exhibition") or el.resolve_default_exhibition()

def _get_code(chat_id: int, args: List[str]) -> str:
    return args[0].strip().upper() if args else _get_current_exhibition(chat_id)

# ---------------------------------------------------------------------------
# Inline Keyboards
# ---------------------------------------------------------------------------

def _main_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Exhibitions", callback_data="menu:exhibitions"), 
         InlineKeyboardButton("Splits", callback_data="menu:splits")],
        [InlineKeyboardButton("Artworks & Sales", callback_data="menu:artworks"), 
         InlineKeyboardButton("Receipts & Expenses", callback_data="menu:expenses")],
        [InlineKeyboardButton("Reports & Export", callback_data="menu:reports"), 
         InlineKeyboardButton("Help & Settings", callback_data="menu:help")],
    ])

def _back_home_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[InlineKeyboardButton("Back to Main Menu", callback_data="menu:home")]])

def _exhibition_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Current Exhibition", callback_data="quick:current"), 
         InlineKeyboardButton("List Portfolio", callback_data="report:portfolio")],
        [InlineKeyboardButton("Add New Exhibition", callback_data="flow_start:new_exhibition"), 
         InlineKeyboardButton("Switch Exhibition", callback_data="flow_start:use_exhibition")],
        [InlineKeyboardButton("Close-Out Checklist", callback_data="report:closeout"), 
         InlineKeyboardButton("Final Readiness", callback_data="report:readiness")],
        [InlineKeyboardButton("Back", callback_data="menu:home")],
    ])

def _split_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("View Current Split", callback_data="report:split")],
        [InlineKeyboardButton("Preset 50/50 Gallery-Artist", callback_data="preset_split:5050")],
        [InlineKeyboardButton("Preset 45/10/45 Gallery-Collaborator-Artist", callback_data="preset_split:451045")],
        [InlineKeyboardButton("Custom Split", callback_data="flow_start:custom_split")],
        [InlineKeyboardButton("Back", callback_data="menu:home")],
    ])

def _artwork_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Register Artwork", callback_data="artwork:register_options"), 
         InlineKeyboardButton("List Artworks", callback_data="report:artworks")],
        [InlineKeyboardButton("Edit Artwork", callback_data="artwork:edit_hint"), 
         InlineKeyboardButton("Inventory Dashboard", callback_data="report:inventory")],
        [InlineKeyboardButton("Record Sale", callback_data="flow_start:record_sale")],
        [InlineKeyboardButton("Back", callback_data="menu:home")],
    ])

def _artwork_register_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Guided Step-by-Step", callback_data="flow_start:add_artwork"),
         InlineKeyboardButton("Bulk Import (Excel)", callback_data="artwork:bulk_import")],
        [InlineKeyboardButton("Back", callback_data="menu:artworks")],
    ])

def _expense_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Add Expense", callback_data="flow_start:receipt"), 
         InlineKeyboardButton("Pending Receipts", callback_data="report:pending")],
        [InlineKeyboardButton("Expense Report", callback_data="report:expenses"), 
         InlineKeyboardButton("Account Heads", callback_data="report:accounts")],
        [InlineKeyboardButton("Set Budget", callback_data="flow_start:set_budget")],
        [InlineKeyboardButton("Back", callback_data="menu:home")],
    ])

def _reports_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Executive Dashboard", callback_data="report:summary"), 
         InlineKeyboardButton("Profit & Loss (P&L)", callback_data="report:pl")],
        [InlineKeyboardButton("Cash Flow Timeline", callback_data="report:cashflow"),
         InlineKeyboardButton("Artist Payouts", callback_data="report:artists")],
        [InlineKeyboardButton("Budget vs Actual", callback_data="report:budget"), 
         InlineKeyboardButton("Data Check", callback_data="report:data_check")],
        [InlineKeyboardButton("Export Excel", callback_data="report:export")],
        [InlineKeyboardButton("Back", callback_data="menu:home")],
    ])

def _help_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Command Guide", callback_data="quick:command_guide"), 
         InlineKeyboardButton("Current Exhibition", callback_data="quick:current")],
        [InlineKeyboardButton("Sheets Status", callback_data="report:sheets_status"), 
         InlineKeyboardButton("Sheets Preview", callback_data="report:sync_preview")],
        [InlineKeyboardButton("Back", callback_data="menu:home")],
    ])

def _pending_keyboard(pending_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Confirm & Post to P&L", callback_data=f"expense:confirm:{pending_id}")],
        [InlineKeyboardButton("Change Account", callback_data=f"expense:account:{pending_id}"),
         InlineKeyboardButton("Correct Amount", callback_data=f"expense:amount:{pending_id}")],
        [InlineKeyboardButton("Ignore Receipt", callback_data=f"expense:ignore:{pending_id}")],
    ])

def _account_keyboard(pending_id: int) -> InlineKeyboardMarkup:
    names = el.account_head_names()
    buttons = []
    for i in range(0, len(names), 2):
        row = [InlineKeyboardButton(names[i][:40], callback_data=f"expense:setacct:{pending_id}:{i}")]
        if i + 1 < len(names):
            row.append(InlineKeyboardButton(names[i + 1][:40], callback_data=f"expense:setacct:{pending_id}:{i + 1}"))
        buttons.append(row)
    buttons.append([InlineKeyboardButton("← Back", callback_data=f"expense:back:{pending_id}")])
    return InlineKeyboardMarkup(buttons)

def _exhibition_picker_keyboard() -> InlineKeyboardMarkup:
    rows = []
    for row in el.list_exhibitions()[:20]:
        rows.append([InlineKeyboardButton(f"{row['code']} — {row['name']}"[:60], callback_data=f"useexh:{row['code']}")])
    rows.append([InlineKeyboardButton("Back", callback_data="menu:exhibitions")])
    return InlineKeyboardMarkup(rows)

async def _send_long_text(message, text: str, parse_mode=None, reply_markup=None) -> None:
    max_len = 3900
    if len(text) <= max_len:
        await message.reply_text(text, parse_mode=parse_mode, reply_markup=reply_markup)
        return
    chunks = [text[i : i + max_len] for i in range(0, len(text), max_len)]
    for idx, chunk in enumerate(chunks):
        await message.reply_text(
            chunk, 
            parse_mode=parse_mode, 
            reply_markup=reply_markup if idx == len(chunks) - 1 else None
        )

# ---------------------------------------------------------------------------
# Core Commands
# ---------------------------------------------------------------------------

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    el.clear_user_flow(chat_id)
    
    code = _get_current_exhibition(chat_id)
    ex = el.get_exhibition(code)
    
    if ex:
        exhibition_line = f"Current working exhibition: *{code} — {ex['name']}*"
    else:
        rows = el.list_exhibitions()
        if rows:
            latest = rows[-1]
            el.set_user_exhibition(chat_id, latest["code"])
            exhibition_line = (
                f"⚠️ Exhibition code resolved from database:\n"
                f"*{latest['code']} — {latest['name']}*\n\n"
                "All exhibitions:\n"
                + "\n".join(f"  • {r['code']} — {r['name']}" for r in rows[:10])
                + "\n\nUse /use <CODE> to switch."
            )
        else:
            exhibition_line = "No exhibitions found. Add an exhibition from the menu to start."

    text = (
        "🎪 *THE SEA ART GALLERY — ExhibitLedger*\n\n"
        "Tap /menu for the guided workflow. All records are in Thai Baht (THB) only.\n\n"
        + exhibition_line
    )
    await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_main_menu_keyboard())

async def menu_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    el.clear_user_flow(chat_id)
    code = _get_current_exhibition(chat_id)
    ex = el.get_exhibition(code)
    label = f"{code} — {ex['name']}" if ex else code
    await update.message.reply_text(f"Main menu. Current exhibition: *{label}*", parse_mode=ParseMode.MARKDOWN, reply_markup=_main_menu_keyboard())

async def cancel_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    el.clear_user_flow(chat_id)
    context.user_data.pop("awaiting_amount_for_pending_id", None)
    await update.message.reply_text("Cancelled the current action.", reply_markup=_main_menu_keyboard())

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    el.clear_user_flow(chat_id)
    text = _command_guide_text()
    await _send_long_text(update.message, text, reply_markup=_help_menu_keyboard())

async def status_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    code = _get_current_exhibition(chat_id)
    try:
        ex = el.get_exhibition(code)
        if not ex:
            rows = el.list_exhibitions()
            if rows:
                names = "\n".join(f"  • {r['code']} — {r['name']}" for r in rows[:10])
                await update.message.reply_text(
                    f"No active exhibition set.\n\nAvailable exhibitions:\n{names}\n\n"
                    "Use /use <CODE> or switch via the menu.",
                    reply_markup=_exhibition_menu_keyboard()
                )
            else:
                await update.message.reply_text("No exhibitions found. Create one first.")
            return

        with el.connect() as conn:
            artworks_total = conn.execute("SELECT COUNT(*) FROM artworks WHERE exhibition_code=?", (code,)).fetchone()[0]
            artworks_sold = conn.execute("SELECT COUNT(*) FROM artworks WHERE exhibition_code=? AND status='sold'", (code,)).fetchone()[0]
            pending_count = conn.execute("SELECT COUNT(*) FROM pending_expenses WHERE exhibition_code=? AND status='pending'", (code,)).fetchone()[0]
            
            # Net P&L logic
            gallery_rev = conn.execute("SELECT SUM(amount_thb) FROM pnl_lines WHERE exhibition_code=? AND section='gallery_revenue'", (code,)).fetchone()[0] or 0.0
            direct_cost = conn.execute("SELECT SUM(amount_thb) FROM pnl_lines WHERE exhibition_code=? AND section='direct_cost'", (code,)).fetchone()[0] or 0.0
            op_exp = conn.execute("SELECT SUM(amount_thb) FROM pnl_lines WHERE exhibition_code=? AND section='operating_expense'", (code,)).fetchone()[0] or 0.0
            overhead = conn.execute("SELECT SUM(amount_thb) FROM pnl_lines WHERE exhibition_code=? AND section='allocated_overhead'", (code,)).fetchone()[0] or 0.0
            net_profit = gallery_rev - direct_cost - op_exp - overhead

        lines = [
            f"📍 *Exhibition:* {code} — {ex['name']} ({ex['status'].upper()})",
            f"🖼️ *Artworks:* {artworks_sold} sold / {artworks_total} total",
            f"🧾 *Pending receipts:* {pending_count} awaiting review",
            f"💵 *Net P&L (Gallery Share):* {el.money(net_profit)}",
        ]
        if pending_count > 0:
            lines.append(f"\n⚠️ {pending_count} receipts need approval. Review with /pending or tap Receipts menu.")
        await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN, reply_markup=_main_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not get status: {exc}")

async def exhibitions(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    rows = el.list_exhibitions()
    if not rows:
        await update.message.reply_text("No exhibitions found. Create one with /new_exhibition CODE Name")
        return
    lines = ["Available exhibitions:\n"]
    for row in rows:
        lines.append(f"• {row['code']} — {row['name']} ({row['status']}; {row['currency']})")
    await update.message.reply_text("\n".join(lines), reply_markup=_exhibition_picker_keyboard())

async def new_exhibition(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    try:
        if len(context.args) < 2:
            raise ValueError("Usage: /new_exhibition <CODE> <NAME>")
        code = el.normalize_code(context.args[0])
        name = " ".join(context.args[1:]).strip()
        row = el.create_exhibition(code, name)
        el.set_user_exhibition(chat_id, row["code"])
        await update.message.reply_text(f"Created exhibition {row['code']} — {row['name']}\nCurrent exhibition set to {row['code']}.", reply_markup=_main_menu_keyboard())
    except Exception as exc:
        logger.exception("Failed to create exhibition")
        await update.message.reply_text(f"Could not create exhibition: {exc}")

async def use_exhibition(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    try:
        if not context.args:
            await update.message.reply_text(f"Current working exhibition: {_get_current_exhibition(chat_id)}", reply_markup=_exhibition_menu_keyboard())
            return
        code = el.normalize_code(context.args[0])
        ex = el.get_exhibition(code)
        if not ex:
            raise ValueError(f"Exhibition not found: {code}")
        el.set_user_exhibition(chat_id, code)
        await update.message.reply_text(f"Current working exhibition set to {code} — {ex['name']}.", reply_markup=_main_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not set current exhibition: {exc}")

async def set_split(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    try:
        code, entries = _parse_split_args(context.args)
        el.set_commission_splits(code, entries)
        el.set_user_exhibition(chat_id, code)
        text = el.format_split_rules_markdown(code)
        await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_split_menu_keyboard())
    except Exception as exc:
        logger.exception("Failed to set split")
        await update.message.reply_text(f"Could not set split: {exc}")

async def split(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    code = _get_code(chat_id, context.args)
    try:
        await update.message.reply_text(el.format_split_rules_markdown(code), parse_mode=ParseMode.MARKDOWN, reply_markup=_split_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not show split for {code}: {exc}")

async def add_artwork_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    try:
        if len(context.args) < 2:
            raise ValueError("Usage: /add_artwork <CODE> <TITLE> | <ARTIST> | <PRICE_THB>")
        code = el.normalize_code(context.args[0])
        rest = " ".join(context.args[1:])
        if "|" in rest:
            parts = _pipe_parts(rest)
            if len(parts) != 3:
                raise ValueError("Use exactly three pipe-separated fields: Title | Artist | PriceTHB")
            title, artist, price_raw = parts
        else:
            if len(context.args) < 4:
                raise ValueError("Usage: /add_artwork <CODE> <TITLE> | <ARTIST> | <PRICE_THB>")
            title, artist, price_raw = context.args[1], context.args[2], context.args[3]
        price = _parse_float(price_raw, "Price")
        row = el.add_artwork(code, title, artist, price)
        el.set_user_exhibition(chat_id, code)
        await update.message.reply_text(
            f"Artwork registered.\nID: #{row['id']}\nExhibition: {row['exhibition_code']}\nTitle: {row['title']}\nArtist: {row['artist']}\nAsking Price: {el.money(row['asking_price_thb'])}",
            reply_markup=_artwork_menu_keyboard(),
        )
    except Exception as exc:
        logger.exception("Failed to add artwork")
        await update.message.reply_text(f"Could not add artwork: {exc}")

async def edit_artwork_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        if len(context.args) < 3:
            raise ValueError(
                "Usage:\n"
                "/edit_artwork <ID> title New Title\n"
                "/edit_artwork <ID> artist New Artist Name\n"
                "/edit_artwork <ID> price 350000"
            )
        artwork_id = int(context.args[0])
        field = context.args[1].lower().strip()
        new_value = " ".join(context.args[2:]).strip()

        if field not in {"title", "artist", "price"}:
            raise ValueError("Field must be title, artist, or price")
        if not new_value:
            raise ValueError("New value cannot be empty.")

        with el.connect() as conn:
            row = conn.execute("SELECT * FROM artworks WHERE id = ?", (artwork_id,)).fetchone()
            if not row:
                raise ValueError(f"Artwork #{artwork_id} not found.")
            row = dict(row)
            if row.get("status") == "sold":
                raise ValueError(f"Artwork #{artwork_id} is already sold and cannot be edited.")

            if field == "price":
                price = _parse_float(new_value, "Price")
                conn.execute("UPDATE artworks SET asking_price_thb = ? WHERE id = ?", (price, artwork_id))
                _log_action_safely("edit_artwork", row["exhibition_code"], f"Updated artwork #{artwork_id} price to {el.money(price)}")
                await update.message.reply_text(f"Artwork #{artwork_id} price updated to {el.money(price)}", reply_markup=_artwork_menu_keyboard())
            elif field == "title":
                conn.execute("UPDATE artworks SET title = ? WHERE id = ?", (new_value, artwork_id))
                _log_action_safely("edit_artwork", row["exhibition_code"], f"Updated artwork #{artwork_id} title to '{new_value}'")
                await update.message.reply_text(f"Artwork #{artwork_id} title updated to '{new_value}'", reply_markup=_artwork_menu_keyboard())
            elif field == "artist":
                conn.execute("UPDATE artworks SET artist = ? WHERE id = ?", (new_value, artwork_id))
                _log_action_safely("edit_artwork", row["exhibition_code"], f"Updated artwork #{artwork_id} artist to '{new_value}'")
                await update.message.reply_text(f"Artwork #{artwork_id} artist updated to '{new_value}'", reply_markup=_artwork_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not edit artwork: {exc}")

async def artworks(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    code = _get_code(chat_id, context.args)
    try:
        await update.message.reply_text(el.format_artworks_markdown(code), parse_mode=ParseMode.MARKDOWN, reply_markup=_artwork_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not list artworks: {exc}")

async def inventory(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    code = _get_code(chat_id, context.args)
    try:
        await update.message.reply_text(el.format_inventory_dashboard_markdown(code), parse_mode=ParseMode.MARKDOWN, reply_markup=_artwork_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not show inventory dashboard: {exc}")

async def sold(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    try:
        if len(context.args) < 2:
            raise ValueError("Usage: /sold <ARTWORK_ID> <ACTUAL_PRICE_THB> [BUYER] [COLLECTED_THB]")
        artwork_id = int(context.args[0])
        actual_price = _parse_float(context.args[1], "Actual sale price")
        buyer_name = None
        collected = None
        if len(context.args) >= 3:
            if _looks_number(context.args[-1]) and len(context.args) >= 4:
                collected = _parse_float(context.args[-1], "Collected amount")
                buyer_name = " ".join(context.args[2:-1]).strip() or None
            else:
                buyer_name = " ".join(context.args[2:]).strip() or None
        result = el.record_sale(artwork_id, actual_price, buyer_name=buyer_name, amount_collected_thb=collected)
        el.set_user_exhibition(chat_id, result["sale"]["exhibition_code"])
        await update.message.reply_text(el.format_sale_markdown(result), parse_mode=ParseMode.MARKDOWN, reply_markup=_artwork_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not record sale: {exc}")

async def receipt(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    try:
        if not context.args:
            raise ValueError("Usage: /receipt <CODE> <AMOUNT_THB> <DESCRIPTION>")
        code, raw_text = _extract_receipt_code_and_text(chat_id, " ".join(context.args))
        pending = el.create_pending_expense(code, raw_text)
        el.set_user_exhibition(chat_id, code)
        await update.message.reply_text(el.format_pending_expense_card(pending), reply_markup=_pending_keyboard(pending["id"]))
    except Exception as exc:
        await update.message.reply_text(f"Could not create receipt: {exc}")

async def pending(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    code = _get_code(chat_id, context.args)
    try:
        await update.message.reply_text(el.format_pending_expenses_markdown(code), reply_markup=_expense_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not show pending receipts: {exc}")

async def expense_report(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    code = _get_code(chat_id, context.args)
    try:
        await update.message.reply_text(el.format_expense_report_markdown(code), reply_markup=_expense_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not generate expense report: {exc}")

async def accounts(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text(el.format_account_heads_markdown(), reply_markup=_expense_menu_keyboard())

async def budget(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    try:
        if len(context.args) >= 3:
            code = el.normalize_code(context.args[0])
            amount = _parse_float(context.args[-1], "Budget amount")
            account = _account_name_from_text(" ".join(context.args[1:-1]))
            el.set_expense_budget(code, account, amount)
            el.set_user_exhibition(chat_id, code)
            await update.message.reply_text(el.format_budget_report_markdown(code), parse_mode=ParseMode.MARKDOWN, reply_markup=_reports_menu_keyboard())
            return
        code = _get_code(chat_id, context.args)
        await update.message.reply_text(el.format_budget_report_markdown(code), parse_mode=ParseMode.MARKDOWN, reply_markup=_reports_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not show budget: {exc}")

async def summary(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    code = _get_code(chat_id, context.args)
    try:
        text = el.format_executive_summary_markdown(code)
        _log_action_safely("telegram_summary", code, "Generated executive summary")
        await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_reports_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not generate summary: {exc}")

async def pl(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    code = _get_code(chat_id, context.args)
    try:
        text = el.format_report_markdown(code)
        _log_action_safely("telegram_pl", code, "Generated P&L")
        await _send_long_text(update.message, text, parse_mode=ParseMode.MARKDOWN, reply_markup=_reports_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not generate P&L: {exc}")

# Cash Flow timeline command
async def cashflow_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    code = _get_code(chat_id, context.args)
    try:
        text = el.format_cash_flow_timeline_markdown(code)
        _log_action_safely("telegram_cashflow", code, "Generated Cash Flow timeline")
        await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_reports_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not generate Cash Flow timeline: {exc}")

# Multi-exhibition Portfolio Dashboard command
async def portfolio_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    try:
        text = el.format_multi_exhibition_dashboard()
        _log_action_safely("telegram_portfolio", None, "Generated Multi-Exhibition Portfolio Dashboard")
        await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_exhibition_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not generate portfolio dashboard: {exc}")

# Close-out checklist command
async def closeout_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    code = _get_code(chat_id, context.args)
    try:
        text = el.format_closeout_status_markdown(code)
        await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_exhibition_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not get close-out checklist: {exc}")

# Close exhibition command
async def close_exhibition_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    code = _get_code(chat_id, context.args)
    try:
        status = el.get_exhibition_closeout_status(code)
        ready = (status["pending_count"] == 0) and (status["unpaid_artists_count"] == 0)
        if not ready:
            await update.message.reply_text(
                f"❌ Cannot close exhibition `{code}` yet.\n"
                f"Please resolve all outstanding checklist items first. Run /close_exhibition_status to view details.",
                reply_markup=_exhibition_menu_keyboard()
            )
            return
        
        el.close_exhibition_in_db(code)
        await update.message.reply_text(f"✅ Exhibition `{code}` has been successfully closed out! Its status is now set to COMPLETED.")
    except Exception as exc:
        await update.message.reply_text(f"Could not close exhibition: {exc}")

async def artist_payouts(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    code = _get_code(chat_id, context.args)
    try:
        text = el.format_artist_payables_markdown(code)
        _log_action_safely("telegram_artist_payouts", code, "Generated artist payouts")
        await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_reports_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not generate artist payouts: {exc}")

async def pay_artist(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    try:
        if len(context.args) < 3:
            raise ValueError(
                "Usage: /pay_artist <CODE> <ARTIST NAME> <AMOUNT THB>\n"
                "Example: /pay_artist CLASH090526 \"Aung Myint\" 50000"
            )
        code = el.normalize_code(context.args[0])
        amount = _parse_float(context.args[-1], "Payment amount")
        artist_name = " ".join(context.args[1:-1]).strip().strip('"').strip("'")
        if not artist_name:
            raise ValueError("Artist name required.")
        if amount <= 0:
            raise ValueError("Payment amount must be greater than zero.")

        with el.connect() as conn:
            row = conn.execute(
                "SELECT * FROM artist_payables WHERE exhibition_code=? AND LOWER(artist)=LOWER(?)",
                (code, artist_name),
            ).fetchone()
            if not row:
                row = conn.execute(
                    "SELECT * FROM artist_payables WHERE exhibition_code=? AND LOWER(artist) LIKE LOWER(?)",
                    (code, f"%{artist_name}%"),
                ).fetchone()
            if not row:
                raise ValueError(f"Artist '{artist_name}' not found in {code}.")
            row = dict(row)
            new_paid = round(row["paid_thb"] + amount, 2)
            new_outstanding = round(row["artist_payable_thb"] - new_paid, 2)
            if new_paid > row["artist_payable_thb"]:
                raise ValueError(f"Payment of {el.money(amount)} exceeds outstanding amount {el.money(row['outstanding_thb'])}.")
            new_status = "Paid" if new_outstanding <= 0 else "Partial"
            conn.execute(
                "UPDATE artist_payables SET paid_thb=?, outstanding_thb=?, status=? WHERE id=?",
                (new_paid, max(new_outstanding, 0), new_status, row["id"]),
            )
            _log_action_safely("pay_artist", code, f"Paid {el.money(amount)} to {row['artist']}")

        lines = [
            f"Payment recorded for {row['artist']}.",
            f"Exhibition: {code}",
            f"This payment: {el.money(amount)}",
            f"Total paid: {el.money(new_paid)} / {el.money(row['artist_payable_thb'])}",
            f"Outstanding: {el.money(max(new_outstanding, 0))}",
            f"Status: {new_status}",
        ]
        await update.message.reply_text("\n".join(lines), reply_markup=_reports_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not record payment: {exc}")

async def readiness(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    code = _get_code(chat_id, context.args)
    try:
        await update.message.reply_text(el.format_readiness_markdown(code), parse_mode=ParseMode.MARKDOWN, reply_markup=_reports_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not check readiness: {exc}")

async def data_check(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    code = _get_code(chat_id, context.args)
    try:
        warnings = el.data_quality_checks(code)
        text = "Data quality check — " + code + "\n\n" + "\n".join(f"• {w}" for w in warnings)
        await update.message.reply_text(text, reply_markup=_reports_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not run data check: {exc}")

async def export(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    code = _get_code(chat_id, context.args)
    try:
        export_dir = os.environ.get("EXPORT_DIR", "./exports")
        file_path = el.export_report_xlsx(code, export_dir)
        with open(file_path, "rb") as f:
            await update.message.reply_document(document=f, filename=Path(file_path).name)
    except Exception as exc:
        await update.message.reply_text(f"Could not export Excel: {exc}")

async def sheets_status(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        await update.message.reply_text(format_sheets_status_markdown(), reply_markup=_help_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not check sheets setup: {exc}")

async def sync_preview(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        await update.message.reply_text(format_sync_preview_markdown(limit_rows=3), reply_markup=_help_menu_keyboard())
    except Exception as exc:
        await update.message.reply_text(f"Could not show sheets preview: {exc}")

async def reseed_shwedagon(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    await update.message.reply_text("Reseeding SHWEDAGON2024... please wait.")
    try:
        db_path = el.db_path()
        raw_conn = sqlite3.connect(db_path)
        try:
            raw_conn.execute("DELETE FROM artist_payables WHERE exhibition_code = 'SHWEDAGON2024'")
            raw_conn.execute("DELETE FROM pnl_lines WHERE exhibition_code = 'SHWEDAGON2024'")
            raw_conn.execute("DELETE FROM exhibitions WHERE code = 'SHWEDAGON2024'")
            raw_conn.commit()
        finally:
            raw_conn.close()
        
        # Import seed logic directly from main.py
        from main import _seed_shwedagon_if_missing
        _seed_shwedagon_if_missing()
        await update.message.reply_text(
            "SHWEDAGON2024 reseeded successfully.\n"
            "26 artists and all P&L lines are now loaded."
        )
    except Exception as exc:
        logger.exception("Reseed failed")
        await update.message.reply_text(f"Reseed failed: {exc}")

# ---------------------------------------------------------------------------
# Menu Navigation Callbacks
# ---------------------------------------------------------------------------

async def menu_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    chat_id = update.effective_chat.id
    
    try:
        await query.answer()
    except BadRequest as e:
        if "Query is too old" in str(e) or "query id is invalid" in str(e).lower():
            logger.warning("Stale callback query ignored")
            return
        raise

    data = query.data or ""
    try:
        if data == "menu:home":
            el.clear_user_flow(chat_id)
            code = _get_current_exhibition(chat_id)
            ex = el.get_exhibition(code)
            label = f"{code} — {ex['name']}" if ex else code
            await query.edit_message_text(f"Main menu. Current exhibition: *{label}*", parse_mode=ParseMode.MARKDOWN, reply_markup=_main_menu_keyboard())
            return
            
        if data == "menu:exhibitions":
            el.clear_user_flow(chat_id)
            code = _get_current_exhibition(chat_id)
            ex = el.get_exhibition(code)
            label = f"{code} — {ex['name']}" if ex else code
            await query.edit_message_text(f"Exhibitions. Current: *{label}*", parse_mode=ParseMode.MARKDOWN, reply_markup=_exhibition_menu_keyboard())
            return
            
        if data == "menu:splits":
            el.clear_user_flow(chat_id)
            code = _get_current_exhibition(chat_id)
            ex = el.get_exhibition(code)
            label = f"{code} — {ex['name']}" if ex else code
            await query.edit_message_text(f"Commission splits. Current exhibition: *{label}*", parse_mode=ParseMode.MARKDOWN, reply_markup=_split_menu_keyboard())
            return
            
        if data == "menu:artworks":
            el.clear_user_flow(chat_id)
            code = _get_current_exhibition(chat_id)
            ex = el.get_exhibition(code)
            label = f"{code} — {ex['name']}" if ex else code
            await query.edit_message_text(f"Artworks and sales. Current exhibition: *{label}*", parse_mode=ParseMode.MARKDOWN, reply_markup=_artwork_menu_keyboard())
            return
            
        if data == "menu:expenses":
            el.clear_user_flow(chat_id)
            code = _get_current_exhibition(chat_id)
            ex = el.get_exhibition(code)
            label = f"{code} — {ex['name']}" if ex else code
            await query.edit_message_text(f"Receipts and expenses. Current exhibition: *{label}*", parse_mode=ParseMode.MARKDOWN, reply_markup=_expense_menu_keyboard())
            return
            
        if data == "menu:reports":
            el.clear_user_flow(chat_id)
            code = _get_current_exhibition(chat_id)
            ex = el.get_exhibition(code)
            label = f"{code} — {ex['name']}" if ex else code
            await query.edit_message_text(f"Reports and export. Current exhibition: *{label}*", parse_mode=ParseMode.MARKDOWN, reply_markup=_reports_menu_keyboard())
            return
            
        if data == "menu:help":
            el.clear_user_flow(chat_id)
            await query.edit_message_text("Help and settings.", reply_markup=_help_menu_keyboard())
            return

        # Handle report callbacks
        if data.startswith("report:"):
            report_name = data.split(":", 1)[1]
            await _handle_report_callback(query, context, report_name)
            return

        if data.startswith("artwork:"):
            action = data.split(":", 1)[1]
            if action == "register_options":
                await query.edit_message_text("Register artwork — choose input method:", reply_markup=_artwork_register_keyboard())
                return
            if action == "bulk_import":
                el.set_user_flow(chat_id, "bulk_artwork")
                try:
                    template_path = generate_artwork_template_xlsx()
                    await query.edit_message_text(
                        "Bulk artwork import:\n\n"
                        "1. Download the template file below\n"
                        "2. Fill in your artworks (delete the 2 sample rows first)\n"
                        "3. Upload the filled file back to this chat\n\n"
                        "Artwork titles with special characters like & work fine."
                    )
                    with open(template_path, "rb") as f:
                        await query.message.reply_document(
                            document=f,
                            filename="artwork_import_template.xlsx",
                            caption="Fill this in and upload it back here to register all artworks at once.",
                        )
                except Exception as exc:
                    await query.message.reply_text(f"Could not generate template: {exc}")
                return
            if action == "edit_hint":
                await query.edit_message_text(
                    "To edit an artwork, send the command:\n\n"
                    "`/edit_artwork <ID> title New Title`\n"
                    "`/edit_artwork <ID> artist New Artist Name`\n"
                    "`/edit_artwork <ID> price 350000`\n\n"
                    "Check artwork IDs with List Artworks.",
                    parse_mode=ParseMode.MARKDOWN,
                    reply_markup=_artwork_menu_keyboard(),
                )
                return

        if data.startswith("useexh:"):
            code = el.normalize_code(data.split(":", 1)[1])
            ex = el.get_exhibition(code)
            if not ex:
                raise ValueError(f"Exhibition not found: {code}")
            el.set_user_exhibition(chat_id, code)
            el.clear_user_flow(chat_id)
            await query.edit_message_text(f"Current working exhibition set to *{code} — {ex['name']}*.", parse_mode=ParseMode.MARKDOWN, reply_markup=_main_menu_keyboard())
            return

        if data.startswith("preset_split:"):
            code = _get_current_exhibition(chat_id)
            preset = data.split(":", 1)[1]
            if preset == "5050":
                entries = [{"party_type": "gallery", "party_name": "Gallery", "percent": 50}, 
                           {"party_type": "artist", "party_name": "Artist", "percent": 50}]
            elif preset == "451045":
                entries = [
                    {"party_type": "gallery", "party_name": "Gallery", "percent": 45},
                    {"party_type": "collaborator", "party_name": "Collaborator", "percent": 10},
                    {"party_type": "artist", "party_name": "Artist", "percent": 45},
                ]
            else:
                raise ValueError("Unknown split preset.")
            el.set_commission_splits(code, entries)
            await query.edit_message_text(el.format_split_rules_markdown(code), parse_mode=ParseMode.MARKDOWN, reply_markup=_split_menu_keyboard())
            return

        if data.startswith("quick:"):
            quick = data.split(":", 1)[1]
            if quick == "current":
                code = _get_current_exhibition(chat_id)
                ex = el.get_exhibition(code)
                text = f"Current working exhibition: *{code}*" + (f" — {ex['name']}" if ex else "")
                await query.edit_message_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_main_menu_keyboard())
                return
            if quick == "command_guide":
                await query.edit_message_text("Command guide sent below.", reply_markup=_help_menu_keyboard())
                await query.message.reply_text(_command_guide_text())
                return

    except Exception as exc:
        logger.exception("Menu callback failed")
        await query.message.reply_text(f"Error: {exc}", reply_markup=_main_menu_keyboard())

# ---------------------------------------------------------------------------
# Passive Expense approval callbacks
# ---------------------------------------------------------------------------

async def expense_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    chat_id = update.effective_chat.id
    
    try:
        await query.answer()
    except BadRequest as e:
        if "Query is too old" in str(e) or "query id is invalid" in str(e).lower():
            logger.warning("Stale expense callback query ignored")
            return
        raise

    try:
        parts = query.data.split(":")
        action = parts[1]
        pending_id = int(parts[2])

        if action == "confirm":
            confirmed = el.confirm_pending_expense(pending_id)
            code = confirmed["exhibition_code"]
            account_head = confirmed["account_head"]
            
            alert = el.check_budget_alert(code, account_head)
            alert_text = f"\n\n{alert}" if alert else ""
            
            await query.edit_message_text(
                f"✅ Expense confirmed & posted to P&L.\n"
                f"Expense ID: #{confirmed['id']}\n"
                f"Exhibition: {code}\n"
                f"Account: {account_head}\n"
                f"Amount: {el.money(confirmed['amount_thb'])}"
                f"{alert_text}",
                parse_mode=ParseMode.MARKDOWN if alert else None,
                reply_markup=_expense_menu_keyboard(),
            )
            return

        if action == "ignore":
            row = el.ignore_pending_expense(pending_id)
            await query.edit_message_text(f"Pending receipt #{row['id']} was ignored.", reply_markup=_expense_menu_keyboard())
            return

        if action == "account":
            await query.edit_message_text(f"Choose account head for pending receipt #{pending_id}:", reply_markup=_account_keyboard(pending_id))
            return

        if action == "setacct":
            account_idx = int(parts[3])
            names = el.account_head_names()
            if account_idx < 0 or account_idx >= len(names):
                raise ValueError("Account index invalid.")
            pending = el.update_pending_account(pending_id, names[account_idx])
            await query.edit_message_text(el.format_pending_expense_card(pending), reply_markup=_pending_keyboard(pending_id))
            return

        if action == "amount":
            context.user_data["awaiting_amount_for_pending_id"] = pending_id
            await query.edit_message_text(f"Send the corrected THB amount for pending receipt #{pending_id}. (e.g. 3500)")
            return

        if action == "back":
            pending = el.get_pending_expense(pending_id)
            if not pending:
                raise ValueError(f"Pending expense #{pending_id} not found.")
            await query.edit_message_text(el.format_pending_expense_card(pending), reply_markup=_pending_keyboard(pending_id))
            return

    except Exception as exc:
        logger.exception("Expense callback failed")
        await query.message.reply_text(f"Could not update receipt: {exc}", reply_markup=_expense_menu_keyboard())

# ---------------------------------------------------------------------------
# Passive receipt capture
# ---------------------------------------------------------------------------

async def handle_text_expense(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    text = update.message.text or ""
    
    if el.parse_amount_thb(text) <= 0:
        await update.message.reply_text(
            "I did not find a valid THB amount. Tap /menu to choose an action, or record an expense by sending, for example:\n"
            f"`{_get_current_exhibition(chat_id)} 3500 coffee and snacks`",
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=_main_menu_keyboard(),
        )
        return

    try:
        code, raw_text = _extract_receipt_code_and_text(chat_id, text)
        pending = el.create_pending_expense(code, raw_text)
        await update.message.reply_text(el.format_pending_expense_card(pending), reply_markup=_pending_keyboard(pending["id"]))
    except Exception as exc:
        logger.exception("Failed to capture text expense")
        await update.message.reply_text(f"Could not capture this expense: {exc}")

# ---------------------------------------------------------------------------
# Command guide text helper
# ---------------------------------------------------------------------------

def _command_guide_text() -> str:
    """Generate a comprehensive help / command guide message."""
    return (
        "📖 THE SEA FINANCE — Command Guide\n\n"
        "🎪 EXHIBITIONS\n"
        "  /start — Start the bot / show main menu\n"
        "  /menu — Open main menu\n"
        "  /status — Current exhibition dashboard\n"
        "  /exhibitions — List all exhibitions\n"
        "  /new_exhibition CODE Name — Create exhibition\n"
        "  /use CODE — Switch active exhibition\n"
        "  /portfolio — Multi-exhibition overview\n"
        "  /close_exhibition_status — Close-out checklist\n"
        "  /close_exhibition — Close an exhibition\n\n"
        "🎨 ARTWORKS\n"
        "  /add_artwork CODE Title | Artist | Price\n"
        "  /edit_artwork ID field NewValue\n"
        "  /artworks — List all artworks\n"
        "  /inventory — Inventory dashboard\n\n"
        "💰 SALES\n"
        "  /sold ARTWORK_ID PRICE [BUYER] [COLLECTED]\n"
        "  /set_split CODE gallery 50 artist 50\n"
        "  /split — View current commission split\n\n"
        "🧾 EXPENSES\n"
        "  /receipt CODE AMOUNT DESCRIPTION\n"
        "  /pending — Review pending receipts\n"
        "  /expense_report — Confirmed expense report\n"
        "  /accounts — List expense categories\n"
        "  /budget CODE CATEGORY AMOUNT — Set budget\n"
        "  📸 Send a photo of a receipt to scan it\n\n"
        "📊 REPORTS\n"
        "  /summary — Executive dashboard\n"
        "  /pl — Full Profit & Loss statement\n"
        "  /cashflow — Cash flow timeline\n"
        "  /artist_payouts — Artist payment status\n"
        "  /pay_artist CODE ARTIST AMOUNT\n"
        "  /readiness — Data readiness check\n"
        "  /data_check — Data quality validation\n"
        "  /export — Download Excel report\n\n"
        "⚙️ OTHER\n"
        "  /cancel — Cancel current action\n"
        "  /help — Show this guide\n"
        "  /sheets_status — Google Sheets integration status\n"
        "  /reseed_shwedagon — Re-import demo data\n\n"
        "💡 TIP: Use /menu for a button-driven interface instead of typing commands."
    )

# ---------------------------------------------------------------------------
# Report callback dispatcher (for menu:report:* callbacks)
# ---------------------------------------------------------------------------

async def _handle_report_callback(query, context: ContextTypes.DEFAULT_TYPE, report_name: str) -> None:
    """Route report:* callback queries to the correct report generator."""
    chat_id = query.from_user.id
    code = _get_current_exhibition(chat_id)

    try:
        if report_name == "summary":
            text = el.format_executive_summary_markdown(code)
            _log_action_safely("telegram_summary", code, "Generated executive summary via menu")
            await query.edit_message_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_reports_menu_keyboard())

        elif report_name == "pl":
            text = el.format_report_markdown(code)
            _log_action_safely("telegram_pl", code, "Generated P&L via menu")
            # P&L can be long; if it exceeds 4096 chars, split across messages
            if len(text) > 4000:
                await query.edit_message_text("Generating P&L report...", reply_markup=None)
                await _send_long_text(query.message, text, parse_mode=ParseMode.MARKDOWN, reply_markup=_reports_menu_keyboard())
            else:
                await query.edit_message_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_reports_menu_keyboard())

        elif report_name == "cashflow":
            text = el.format_cash_flow_timeline_markdown(code)
            _log_action_safely("telegram_cashflow", code, "Generated Cash Flow via menu")
            await query.edit_message_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_reports_menu_keyboard())

        elif report_name == "artists":
            text = el.format_artist_payables_markdown(code)
            _log_action_safely("telegram_artist_payouts", code, "Generated artist payouts via menu")
            await query.edit_message_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_reports_menu_keyboard())

        elif report_name == "budget":
            text = el.format_budget_report_markdown(code)
            await query.edit_message_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_reports_menu_keyboard())

        elif report_name == "data_check":
            warnings = el.data_quality_checks(code)
            text = "Data quality check — " + code + "\n\n" + "\n".join(f"• {w}" for w in warnings)
            await query.edit_message_text(text, reply_markup=_reports_menu_keyboard())

        elif report_name == "export":
            await query.edit_message_text("Generating Excel report...")
            export_dir = os.environ.get("EXPORT_DIR", "./exports")
            file_path = el.export_report_xlsx(code, export_dir)
            with open(file_path, "rb") as f:
                await query.message.reply_document(document=f, filename=Path(file_path).name)

        elif report_name == "artworks":
            text = el.format_artworks_markdown(code)
            await query.edit_message_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_artwork_menu_keyboard())

        elif report_name == "inventory":
            text = el.format_inventory_dashboard_markdown(code)
            await query.edit_message_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_artwork_menu_keyboard())

        elif report_name == "pending":
            text = el.format_pending_expenses_markdown(code)
            await query.edit_message_text(text, reply_markup=_expense_menu_keyboard())

        elif report_name == "expenses":
            text = el.format_expense_report_markdown(code)
            await query.edit_message_text(text, reply_markup=_expense_menu_keyboard())

        elif report_name == "accounts":
            text = el.format_account_heads_markdown()
            await query.edit_message_text(text, reply_markup=_expense_menu_keyboard())

        elif report_name == "split":
            text = el.format_split_rules_markdown(code)
            await query.edit_message_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_split_menu_keyboard())

        elif report_name == "portfolio":
            text = el.format_multi_exhibition_dashboard()
            _log_action_safely("telegram_portfolio", None, "Generated Portfolio Dashboard via menu")
            await query.edit_message_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_exhibition_menu_keyboard())

        elif report_name == "closeout":
            text = el.format_closeout_status_markdown(code)
            await query.edit_message_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_exhibition_menu_keyboard())

        elif report_name == "readiness":
            text = el.format_readiness_markdown(code)
            await query.edit_message_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=_exhibition_menu_keyboard())

        elif report_name == "sheets_status":
            text = format_sheets_status_markdown()
            await query.edit_message_text(text, reply_markup=_help_menu_keyboard())

        elif report_name == "sync_preview":
            text = format_sync_preview_markdown(limit_rows=3)
            await query.edit_message_text(text, reply_markup=_help_menu_keyboard())

        else:
            await query.edit_message_text(f"Unknown report: {report_name}", reply_markup=_reports_menu_keyboard())

    except Exception as exc:
        logger.exception(f"Report callback failed: {report_name}")
        await query.message.reply_text(f"Could not generate report: {exc}", reply_markup=_reports_menu_keyboard())

# ---------------------------------------------------------------------------
# Guided Flow Definitions
# ---------------------------------------------------------------------------
#
# Each flow is a multi-step conversation. The flow state is persisted to the
# database so it survives bot restarts. Steps are defined as dictionaries:
#   { "prompt": str, "field": str, "type": "text"|"number"|"choice", "choices": [...] }
#

GUIDED_FLOWS = {
    "new_exhibition": {
        "title": "Create New Exhibition",
        "steps": [
            {"prompt": "What's the exhibition CODE? (e.g. CLASH090526, uppercase, no spaces)", "field": "code", "type": "text"},
            {"prompt": "What's the exhibition name?", "field": "name", "type": "text"},
        ],
    },
    "use_exhibition": {
        "title": "Switch Exhibition",
        "steps": [],  # Handled specially with a picker keyboard
    },
    "add_artwork": {
        "title": "Register Artwork",
        "steps": [
            {"prompt": "What's the artwork title?", "field": "title", "type": "text"},
            {"prompt": "Who's the artist?", "field": "artist", "type": "text"},
            {"prompt": "What's the asking price in THB? (number only, e.g. 150000)", "field": "price", "type": "number"},
        ],
    },
    "record_sale": {
        "title": "Record a Sale",
        "steps": [
            {"prompt": "What's the artwork ID number? (check with List Artworks)", "field": "artwork_id", "type": "number"},
            {"prompt": "What was the actual selling price in THB?", "field": "price", "type": "number"},
            {"prompt": "Buyer name? (or type 'skip' to leave blank)", "field": "buyer", "type": "text"},
        ],
    },
    "receipt": {
        "title": "Add Expense / Receipt",
        "steps": [
            {"prompt": "Enter the expense amount in THB (number only, e.g. 5000):", "field": "amount", "type": "number"},
            {"prompt": "What was this expense for? (brief description, e.g. 'printing exhibition catalogs')", "field": "description", "type": "text"},
        ],
    },
    "custom_split": {
        "title": "Custom Commission Split",
        "steps": [
            {
                "prompt": (
                    "Enter the commission split as space-separated tokens.\n\n"
                    "Format: gallery PERCENT artist PERCENT\n"
                    "With collaborator: gallery 40 collaborator Rotary 10 artist 50\n\n"
                    "Example: gallery 50 artist 50"
                ),
                "field": "split_text",
                "type": "text",
            },
        ],
    },
    "set_budget": {
        "title": "Set Expense Budget",
        "steps": [
            {"prompt": "Which expense category? (e.g. Printing, Venue Rental, Transport)\nSee /accounts for full list.", "field": "account", "type": "text"},
            {"prompt": "Budget amount in THB?", "field": "amount", "type": "number"},
        ],
    },
}


async def flow_start_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle flow_start:* callbacks — begin a guided multi-step flow."""
    query = update.callback_query
    chat_id = update.effective_chat.id

    try:
        await query.answer()
    except BadRequest as e:
        if "Query is too old" in str(e) or "query id is invalid" in str(e).lower():
            return
        raise

    data = query.data or ""
    flow_name = data.split(":", 1)[1] if ":" in data else ""

    try:
        # Special case: use_exhibition shows a picker instead of a step-based flow
        if flow_name == "use_exhibition":
            el.clear_user_flow(chat_id)
            await query.edit_message_text(
                "Select an exhibition to switch to:",
                reply_markup=_exhibition_picker_keyboard(),
            )
            return

        flow_def = GUIDED_FLOWS.get(flow_name)
        if not flow_def:
            await query.edit_message_text(
                f"Unknown guided flow: {flow_name}. Tap /menu to go back.",
                reply_markup=_main_menu_keyboard(),
            )
            return

        if not flow_def["steps"]:
            await query.edit_message_text(
                "This flow is not yet configured. Use the command-line equivalent instead.",
                reply_markup=_main_menu_keyboard(),
            )
            return

        # Initialize flow state
        el.set_user_flow(chat_id, flow_name, flow_step=0, flow_data={})
        first_step = flow_def["steps"][0]
        await query.edit_message_text(
            f"📝 *{flow_def['title']}* — Step 1/{len(flow_def['steps'])}\n\n{first_step['prompt']}",
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton("❌ Cancel", callback_data="flow_cb:cancel")]]
            ),
        )

    except Exception as exc:
        logger.exception("Flow start failed")
        await query.message.reply_text(f"Could not start flow: {exc}", reply_markup=_main_menu_keyboard())


async def flow_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle flow_cb:* callbacks (e.g. cancel, account selection within a flow)."""
    query = update.callback_query
    chat_id = update.effective_chat.id

    try:
        await query.answer()
    except BadRequest as e:
        if "Query is too old" in str(e) or "query id is invalid" in str(e).lower():
            return
        raise

    data = query.data or ""
    action = data.split(":", 1)[1] if ":" in data else ""

    try:
        if action == "cancel":
            el.clear_user_flow(chat_id)
            await query.edit_message_text("Flow cancelled.", reply_markup=_main_menu_keyboard())
            return

        # Future: handle in-flow choice selections (e.g. account head picker within receipt flow)
        await query.edit_message_text("Unknown flow action. Tap /menu to restart.", reply_markup=_main_menu_keyboard())

    except Exception as exc:
        logger.exception("Flow callback failed")
        await query.message.reply_text(f"Error: {exc}", reply_markup=_main_menu_keyboard())


# ---------------------------------------------------------------------------
# Master text input handler
# ---------------------------------------------------------------------------

async def handle_user_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Main router for all plain-text messages (non-command, non-photo, non-document).
    
    Priority:
    1. Check if user is in the middle of a guided flow → advance the flow
    2. Check if user is correcting an expense amount → apply correction
    3. Fall back to passive expense capture (handle_text_expense)
    """
    chat_id = update.effective_chat.id
    text = (update.message.text or "").strip()

    if not text:
        return

    # ----- 1. Check for active guided flow -----
    state = el.get_user_state(chat_id)
    active_flow = state.get("active_flow")

    if active_flow and active_flow in GUIDED_FLOWS:
        await _handle_flow_step(update, context, active_flow, state)
        return

    # ----- 2. Check for pending amount correction -----
    awaiting_id = context.user_data.get("awaiting_amount_for_pending_id")
    if awaiting_id:
        try:
            amount = _parse_float(text, "Amount")
            if amount <= 0:
                raise ValueError("Amount must be greater than zero.")
            pending = el.update_pending_amount(awaiting_id, amount)
            context.user_data.pop("awaiting_amount_for_pending_id", None)
            await update.message.reply_text(
                el.format_pending_expense_card(pending),
                reply_markup=_pending_keyboard(pending["id"]),
            )
        except Exception as exc:
            await update.message.reply_text(f"Invalid amount: {exc}\nSend a number, e.g. 3500")
        return

    # ----- 3. Fall back to passive expense capture -----
    await handle_text_expense(update, context)


async def _handle_flow_step(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    flow_name: str,
    state: dict,
) -> None:
    """Process one step of a guided flow, advancing state through the DB."""
    chat_id = update.effective_chat.id
    text = (update.message.text or "").strip()
    flow_def = GUIDED_FLOWS[flow_name]
    steps = flow_def["steps"]
    current_step = state.get("flow_step", 0)

    # Load existing flow data
    try:
        flow_data = json.loads(state.get("flow_data") or "{}")
    except (json.JSONDecodeError, TypeError):
        flow_data = {}

    if current_step >= len(steps):
        # Shouldn't happen, but clear flow and let user restart
        el.clear_user_flow(chat_id)
        await update.message.reply_text("Flow completed. Tap /menu to continue.", reply_markup=_main_menu_keyboard())
        return

    step_def = steps[current_step]
    field = step_def["field"]
    step_type = step_def.get("type", "text")

    # Validate input
    try:
        if step_type == "number":
            value = _parse_float(text, step_def.get("prompt", "Value"))
            if value <= 0:
                raise ValueError("Must be a positive number.")
            flow_data[field] = value
        else:
            if not text:
                raise ValueError("Please enter a value.")
            flow_data[field] = text
    except ValueError as exc:
        await update.message.reply_text(
            f"⚠️ {exc}\n\n{step_def['prompt']}",
            reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton("❌ Cancel", callback_data="flow_cb:cancel")]]
            ),
        )
        return

    next_step = current_step + 1

    if next_step < len(steps):
        # Save progress and prompt next step
        el.set_user_flow(chat_id, flow_name, flow_step=next_step, flow_data=flow_data)
        next_step_def = steps[next_step]
        await update.message.reply_text(
            f"📝 *{flow_def['title']}* — Step {next_step + 1}/{len(steps)}\n\n{next_step_def['prompt']}",
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton("❌ Cancel", callback_data="flow_cb:cancel")]]
            ),
        )
    else:
        # All steps collected — execute the flow action
        el.clear_user_flow(chat_id)
        await _execute_flow(update, context, flow_name, flow_data)


async def _execute_flow(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    flow_name: str,
    flow_data: dict,
) -> None:
    """Execute the final action for a completed guided flow."""
    chat_id = update.effective_chat.id
    code = _get_current_exhibition(chat_id)

    try:
        if flow_name == "new_exhibition":
            new_code = el.normalize_code(flow_data["code"])
            name = flow_data["name"]
            row = el.create_exhibition(new_code, name)
            el.set_user_exhibition(chat_id, row["code"])
            await update.message.reply_text(
                f"✅ Exhibition created!\n\n"
                f"Code: {row['code']}\n"
                f"Name: {row['name']}\n\n"
                f"Current exhibition switched to {row['code']}.",
                reply_markup=_main_menu_keyboard(),
            )

        elif flow_name == "add_artwork":
            title = flow_data["title"]
            artist = flow_data["artist"]
            price = flow_data["price"]
            row = el.add_artwork(code, title, artist, price)
            _log_action_safely("add_artwork", code, f"Registered '{title}' by {artist} at {el.money(price)}")
            await update.message.reply_text(
                f"✅ Artwork registered!\n\n"
                f"ID: #{row['id']}\n"
                f"Exhibition: {row['exhibition_code']}\n"
                f"Title: {row['title']}\n"
                f"Artist: {row['artist']}\n"
                f"Asking Price: {el.money(row['asking_price_thb'])}",
                reply_markup=_artwork_menu_keyboard(),
            )

        elif flow_name == "record_sale":
            artwork_id = int(flow_data["artwork_id"])
            price = flow_data["price"]
            buyer = flow_data.get("buyer", "").strip()
            if buyer.lower() in ("skip", "", "none", "-"):
                buyer = None
            result = el.record_sale(artwork_id, price, buyer_name=buyer)
            el.set_user_exhibition(chat_id, result["sale"]["exhibition_code"])
            _log_action_safely("record_sale", result["sale"]["exhibition_code"], f"Sold artwork #{artwork_id}")
            await update.message.reply_text(
                el.format_sale_markdown(result),
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=_artwork_menu_keyboard(),
            )

        elif flow_name == "receipt":
            amount = flow_data["amount"]
            description = flow_data["description"]
            raw_text = f"{amount} {description}"
            pending = el.create_pending_expense(code, raw_text)
            _log_action_safely("add_expense", code, f"Added pending expense: {el.money(amount)} — {description}")
            await update.message.reply_text(
                el.format_pending_expense_card(pending),
                reply_markup=_pending_keyboard(pending["id"]),
            )

        elif flow_name == "custom_split":
            split_text = flow_data["split_text"]
            tokens = split_text.strip().split()
            _, entries = _parse_split_tokens(code, tokens)
            el.set_commission_splits(code, entries)
            _log_action_safely("set_split", code, f"Custom split set via guided flow")
            await update.message.reply_text(
                el.format_split_rules_markdown(code),
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=_split_menu_keyboard(),
            )

        elif flow_name == "set_budget":
            account = _account_name_from_text(flow_data["account"])
            amount = flow_data["amount"]
            el.set_expense_budget(code, account, amount)
            _log_action_safely("set_budget", code, f"Budget set: {account} = {el.money(amount)}")
            await update.message.reply_text(
                el.format_budget_report_markdown(code),
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=_reports_menu_keyboard(),
            )

        else:
            await update.message.reply_text(
                f"Flow '{flow_name}' completed but no action handler exists.",
                reply_markup=_main_menu_keyboard(),
            )

    except Exception as exc:
        logger.exception(f"Flow execution failed: {flow_name}")
        await update.message.reply_text(f"Could not complete {flow_name}: {exc}", reply_markup=_main_menu_keyboard())

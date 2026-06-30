import base64
import io
import logging
import os
import re
import sqlite3
import shutil
import asyncio
from datetime import datetime
from pathlib import Path
from typing import List, Tuple
from contextlib import asynccontextmanager

import pytesseract
from PIL import Image
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.constants import ParseMode
from telegram.error import BadRequest, Conflict, Forbidden, NetworkError, TelegramError, TimedOut
from telegram.ext import Application, CallbackQueryHandler, CommandHandler, ContextTypes, MessageHandler, filters

from fastapi import FastAPI, Request, Response
import uvicorn

import exhibitledger as el
import handlers

# Optional OpenAI client for enhanced receipt processing
ai_client = None
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")
if OPENAI_API_KEY and OPENAI_API_KEY.strip():
    try:
        from openai import OpenAI
        ai_client = OpenAI()
        logging.info("OpenAI client initialized for enhanced receipt scanning.")
    except Exception as e:
        logging.warning(f"Could not initialize OpenAI client: {e}. Using OCR fallback.")

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=os.environ.get("LOG_LEVEL", "INFO"),
)
logger = logging.getLogger(__name__)
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("telegram.ext").setLevel(logging.INFO)

# Helper to get current exhibition for document upload
def _get_current_exhibition(chat_id: int) -> str:
    state = el.get_user_state(chat_id)
    return state.get("current_exhibition") or el.resolve_default_exhibition()

# ---------------------------------------------------------------------------
# Document & Image Handlers
# ---------------------------------------------------------------------------

async def handle_document_upload(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle Excel file uploads for bulk artwork import."""
    doc = update.message.document
    if not doc:
        return

    fname = doc.file_name or ""
    if not fname.lower().endswith((".xlsx", ".xls")):
        chat_id = update.effective_chat.id
        state = el.get_user_state(chat_id)
        if state.get("active_flow") == "bulk_artwork":
            await update.message.reply_text(
                "Please upload an Excel (.xlsx) file.\n"
                "Use the template from Register Artwork → Bulk Import, fill it in, and upload it back."
            )
        return

    chat_id = update.effective_chat.id
    code = _get_current_exhibition(chat_id)
    ex = el.get_exhibition(code)
    
    # Clear flow
    el.clear_user_flow(chat_id)

    await update.message.reply_text(
        f"Received {fname}.\nProcessing artwork import for {code}"
        + (f" — {ex['name']}" if ex else "")
        + "..."
    )

    try:
        from openpyxl import load_workbook

        file_obj = await doc.get_file()
        raw = await file_obj.download_as_bytearray()
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        registered = []
        duplicates = []
        skipped = []

        with el.connect() as conn:
            existing_titles = {
                row[0].lower()
                for row in conn.execute(
                    "SELECT title FROM artworks WHERE exhibition_code=?", (code,)
                ).fetchall()
            }

        for i, row in enumerate(rows, start=2):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            title = str(row[0]).strip() if row[0] is not None else ""
            artist = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            price_raw = row[2] if len(row) > 2 else None

            if "delete this row" in title.lower() or "delete this row" in artist.lower() or "sample title" in title.lower():
                continue

            if not title or not artist:
                skipped.append(f"Row {i}: missing title or artist name")
                continue

            if title.lower() in existing_titles:
                duplicates.append(title)
                continue

            try:
                price = float(str(price_raw).replace(",", "").strip())
                if price <= 0:
                    raise ValueError("must be greater than zero")
            except Exception:
                skipped.append(f"Row {i} — {title}: invalid price \'{price_raw}\'")
                continue

            try:
                artwork_row = el.add_artwork(code, title, artist, price)
                registered.append(f"#{artwork_row['id']}  {title}  |  {artist}  |  {el.money(price)}")
                existing_titles.add(title.lower())
            except Exception as exc:
                skipped.append(f"Row {i} — {title}: {exc}")

        if not registered and not duplicates and not skipped:
            await update.message.reply_text(
                "The file appears empty or has only a header row.\n\n"
                "Make sure you filled in artworks below the header row.",
                reply_markup=handlers._artwork_menu_keyboard(),
            )
            return

        lines = [f"Artwork import — {code}", ""]
        if registered:
            lines.append(f"Registered {len(registered)} new artwork(s):")
            lines.extend(f"  ✅ {r}" for r in registered)
        if duplicates:
            if registered:
                lines.append("")
            lines.append(f"Already registered — skipped {len(duplicates)}:")
            lines.extend(f"  ⏭ {d}" for d in duplicates)
            lines.append("  Use /edit_artwork to update existing artworks.")
        if skipped:
            lines.append("")
            lines.append(f"Could not import {len(skipped)} row(s):")
            lines.extend(f"  ⚠️ {s}" for s in skipped)
        if registered:
            lines += ["", "Tap List Artworks to confirm, or /export for the full report."]
        elif duplicates and not registered:
            lines += ["", "All artworks from this file are already registered."]

        await update.message.reply_text("\n".join(lines), reply_markup=handlers._artwork_menu_keyboard())

    except Exception as exc:
        logger.exception("Bulk artwork import failed")
        await update.message.reply_text(
            f"Could not read the file: {exc}\n\n"
            "Make sure it is a valid .xlsx in the artwork import template format.\n"
            "Tap Register Artwork → Bulk Import to get a fresh template."
        )

async def _process_receipt_image(photo_file) -> str:
    """Extract text from a photo using OCR or AI."""
    image_bytes = await photo_file.download_as_bytearray()
    image = Image.open(io.BytesIO(image_bytes))

    # Try AI first
    if ai_client:
        try:
            buffered = io.BytesIO()
            image.save(buffered, format="JPEG")
            base64_image = base64.b64encode(buffered.getvalue()).decode("utf-8")

            response = ai_client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": "Extract all text from this receipt. Focus on the total amount and items."},
                            {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}},
                        ],
                    }
                ],
                max_tokens=300,
            )
            ai_text = response.choices[0].message.content
            if ai_text:
                return ai_text
        except Exception as e:
            logger.warning(f"AI receipt extraction failed: {e}. Falling back to Tesseract.")

    # Fallback to Tesseract OCR
    try:
        return pytesseract.image_to_string(image)
    except Exception as e:
        logger.error(f"Tesseract OCR failed: {e}")
        return ""

async def handle_photo_expense(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        chat_id = update.effective_chat.id
        status_msg = await update.message.reply_text("Scanning receipt... 🔍")

        caption = update.message.caption or ""
        photo = update.message.photo[-1]
        photo_file = await photo.get_file()

        ocr_text = await _process_receipt_image(photo_file)
        combined_text = f"{caption}\n\n{ocr_text}".strip()

        from handlers import _extract_receipt_code_and_text, _pending_keyboard
        code, raw_text = _extract_receipt_code_and_text(chat_id, combined_text)
        photo_file_id = photo.file_id

        pending = el.create_pending_expense(code, raw_text, photo_file_id=photo_file_id)

        try:
            await status_msg.delete()
        except Exception:
            pass
            
        await update.message.reply_text(
            el.format_pending_expense_card(pending),
            reply_markup=_pending_keyboard(pending["id"])
        )
    except Exception as exc:
        logger.exception("Failed to capture photo receipt")
        await update.message.reply_text(f"Could not capture this receipt photo: {exc}")

# ---------------------------------------------------------------------------
# Error Handling
# ---------------------------------------------------------------------------

async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    logger.error("Exception while handling update:", exc_info=context.error)

    if isinstance(context.error, TimedOut):
        return
    if isinstance(context.error, NetworkError):
        return
    if isinstance(context.error, Conflict):
        logger.warning("Conflict: Another instance is polling.")
        return
    if isinstance(context.error, Forbidden):
        return

    if isinstance(update, Update) and update.effective_message:
        try:
            await update.effective_message.reply_text(
                "Something went wrong. Please try again or tap /menu to restart."
            )
        except Exception:
            pass

# ---------------------------------------------------------------------------
# Automated Backups
# ---------------------------------------------------------------------------

async def run_backup(application: Application) -> None:
    """Creates a backup copy of the database and sends it to active chats."""
    db_path = el.db_path()
    if not os.path.exists(db_path):
        logger.warning(f"Database not found at {db_path}, skipping backup.")
        return
        
    now = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    backup_filename = f"exhibitledger_backup_{now}.db"
    
    export_dir = os.environ.get("EXPORT_DIR", "./exports")
    os.makedirs(export_dir, exist_ok=True)
    backup_path = os.path.join(export_dir, backup_filename)
    
    try:
        shutil.copy2(db_path, backup_path)
        logger.info(f"Database backed up locally to {backup_path}")
        
        target_chats = set()
        
        # 1. From environment variable
        backup_chat_id = os.environ.get("BACKUP_CHAT_ID")
        if backup_chat_id:
            try:
                target_chats.add(int(backup_chat_id))
            except ValueError:
                logger.warning(f"Invalid BACKUP_CHAT_ID: {backup_chat_id}")
                
        # 2. Fallback to active chat IDs in user states
        if not target_chats:
            try:
                with el.connect() as conn:
                    rows = conn.execute("SELECT chat_id FROM user_states").fetchall()
                    for r in rows:
                        target_chats.add(r[0])
            except Exception as e:
                logger.warning(f"Could not retrieve chat IDs from user_states: {e}")
                
        if not target_chats:
            logger.warning("No chat IDs found for sending backup. Set BACKUP_CHAT_ID in environment.")
            return
            
        try:
            with el.connect() as conn:
                exh_count = conn.execute("SELECT COUNT(*) FROM exhibitions").fetchone()[0]
                art_count = conn.execute("SELECT COUNT(*) FROM artworks").fetchone()[0]
                sales_count = conn.execute("SELECT COUNT(*) FROM artwork_sales").fetchone()[0]
        except Exception:
            exh_count = art_count = sales_count = 0
            
        caption = (
            f"🗄️ *ExhibitLedger Daily Backup*\n"
            f"📅 Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"📊 Portfolio: {exh_count} exhibitions, {art_count} artworks, {sales_count} sales"
        )
        
        for chat_id in target_chats:
            try:
                with open(backup_path, "rb") as f:
                    await application.bot.send_document(
                        chat_id=chat_id,
                        document=f,
                        filename=backup_filename,
                        caption=caption,
                        parse_mode=ParseMode.MARKDOWN
                    )
                logger.info(f"Backup sent successfully to chat_id {chat_id}")
            except Exception as send_err:
                logger.error(f"Failed to send backup to chat_id {chat_id}: {send_err}")
                
        try:
            os.remove(backup_path)
        except Exception:
            pass
            
    except Exception as e:
        logger.error(f"Backup failed: {e}")

async def backup_scheduler(application: Application) -> None:
    """Async loop running daily backups in the background."""
    logger.info("Database backup scheduler started.")
    # Wait 10 seconds on startup to verify backup functionality
    await asyncio.sleep(10)
    await run_backup(application)
    
    while True:
        await asyncio.sleep(86400) # 24 hours
        await run_backup(application)

# Helper function to run backup scheduler
async def post_init(application: Application) -> None:
    asyncio.create_task(backup_scheduler(application))

# ---------------------------------------------------------------------------
# Bot Seeding & Setup
# ---------------------------------------------------------------------------

def _seed_shwedagon_if_missing() -> None:
    try:
        el.seed_shwedagon_if_missing()
    except Exception as e:
        logger.warning(f"Could not seed SHWEDAGON2024: {e}")

def build_application() -> Application:
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    if not token:
        raise RuntimeError("TELEGRAM_BOT_TOKEN is not set.")
    el.init_db()

    # Auto-seed SHWEDAGON2024 on startup if missing or incomplete
    try:
        db_path = el.db_path()
        _raw = sqlite3.connect(db_path)
        try:
            _ex = _raw.execute("SELECT code FROM exhibitions WHERE code='SHWEDAGON2024'").fetchone()
            _artist_count = _raw.execute(
                "SELECT COUNT(*) FROM artist_payables WHERE exhibition_code='SHWEDAGON2024'"
            ).fetchone()[0]
            if not _ex or _artist_count < 5:
                _raw.execute("DELETE FROM artist_payables WHERE exhibition_code='SHWEDAGON2024'")
                _raw.execute("DELETE FROM pnl_lines WHERE exhibition_code='SHWEDAGON2024'")
                _raw.execute("DELETE FROM exhibitions WHERE code='SHWEDAGON2024'")
                _raw.commit()
                _needs_seed = True
            else:
                _needs_seed = False
        finally:
            _raw.close()

        if _needs_seed:
            _seed_shwedagon_if_missing()
            logger.info("SHWEDAGON2024 seeded/reseeded on startup.")
    except Exception as _seed_err:
        logger.warning("Could not check/seed SHWEDAGON2024 on startup: %s", _seed_err)

    application = Application.builder().token(token).post_init(post_init).build()
    
    # Register command handlers
    application.add_handler(CommandHandler("start", handlers.start))
    application.add_handler(CommandHandler("menu", handlers.menu_command))
    application.add_handler(CommandHandler("status", handlers.status_command))
    application.add_handler(CommandHandler("cancel", handlers.cancel_command))
    application.add_handler(CommandHandler("help", handlers.help_command))
    application.add_handler(CommandHandler("exhibitions", handlers.exhibitions))
    application.add_handler(CommandHandler("new_exhibition", handlers.new_exhibition))
    application.add_handler(CommandHandler("use", handlers.use_exhibition))
    application.add_handler(CommandHandler("set_split", handlers.set_split))
    application.add_handler(CommandHandler("split", handlers.split))
    application.add_handler(CommandHandler("add_artwork", handlers.add_artwork_command))
    application.add_handler(CommandHandler("edit_artwork", handlers.edit_artwork_command))
    application.add_handler(CommandHandler("artworks", handlers.artworks))
    application.add_handler(CommandHandler("inventory", handlers.inventory))
    application.add_handler(CommandHandler("sold", handlers.sold))
    application.add_handler(CommandHandler("receipt", handlers.receipt))
    application.add_handler(CommandHandler("pending", handlers.pending))
    application.add_handler(CommandHandler("expense_report", handlers.expense_report))
    application.add_handler(CommandHandler("accounts", handlers.accounts))
    application.add_handler(CommandHandler("budget", handlers.budget))
    application.add_handler(CommandHandler("summary", handlers.summary))
    application.add_handler(CommandHandler("pl", handlers.pl))
    application.add_handler(CommandHandler("cashflow", handlers.cashflow_command))
    application.add_handler(CommandHandler("portfolio", handlers.portfolio_command))
    application.add_handler(CommandHandler("artist_payouts", handlers.artist_payouts))
    application.add_handler(CommandHandler("pay_artist", handlers.pay_artist))
    application.add_handler(CommandHandler("close_exhibition_status", handlers.closeout_command))
    application.add_handler(CommandHandler("close_exhibition", handlers.close_exhibition_command))
    application.add_handler(CommandHandler("readiness", handlers.readiness))
    application.add_handler(CommandHandler("data_check", handlers.data_check))
    application.add_handler(CommandHandler("export", handlers.export))
    application.add_handler(CommandHandler("sheets_status", handlers.sheets_status))
    application.add_handler(CommandHandler("sync_preview", handlers.sync_preview))
    application.add_handler(CommandHandler("reseed_shwedagon", handlers.reseed_shwedagon))
    
    # Register callback query handlers
    application.add_handler(CallbackQueryHandler(handlers.expense_callback, pattern=r"^expense:"))
    application.add_handler(CallbackQueryHandler(handlers.flow_start_callback, pattern=r"^flow_start:"))
    application.add_handler(CallbackQueryHandler(handlers.flow_callback, pattern=r"^flow_cb:"))
    application.add_handler(CallbackQueryHandler(handlers.menu_callback, pattern=r"^(menu:|artwork:|preset_split:|useexh:|quick:|report:)"))
    
    # Register message handlers
    application.add_handler(MessageHandler(filters.Document.ALL, handle_document_upload))
    application.add_handler(MessageHandler(filters.PHOTO, handle_photo_expense))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handlers.handle_user_input))

    application.add_error_handler(error_handler)

    return application

# ---------------------------------------------------------------------------
# FastAPI & Lifespan Setup for Webhooks
# ---------------------------------------------------------------------------

bot_app = None

@asynccontextmanager
async def lifespan(app: FastAPI):
    global bot_app
    bot_app = build_application()
    
    await bot_app.initialize()
    await bot_app.start()
    
    webhook_url = os.environ.get("RENDER_EXTERNAL_URL") or os.environ.get("WEBHOOK_URL")
    if webhook_url:
        token = os.environ.get("TELEGRAM_BOT_TOKEN")
        if not webhook_url.startswith("http"):
            webhook_url = f"https://{webhook_url}"
        await bot_app.bot.set_webhook(url=f"{webhook_url}/telegram")
        logger.info(f"Telegram webhook set to: {webhook_url}/telegram")
    else:
        logger.warning("RENDER_EXTERNAL_URL or WEBHOOK_URL not set. Running webhook server without registering webhook.")
        
    yield
    
    if bot_app:
        await bot_app.stop()
        await bot_app.shutdown()

app = FastAPI(lifespan=lifespan)

@app.post("/telegram")
async def webhook_handler(request: Request):
    if bot_app:
        data = await request.json()
        update = Update.de_json(data, bot_app.bot)
        await bot_app.update_queue.put(update)
    return Response(status_code=200)

@app.get("/health")
@app.get("/")
def health_check():
    return {"status": "ok", "bot": "running"}

# ---------------------------------------------------------------------------
# Main Entry Point
# ---------------------------------------------------------------------------

def main() -> None:
    port = int(os.environ.get("PORT", "10000"))
    
    # Check if we should run in polling mode (for local testing/fallback)
    polling_mode = os.environ.get("POLLING", "False").lower() in ("true", "1")
    webhook_url = os.environ.get("RENDER_EXTERNAL_URL") or os.environ.get("WEBHOOK_URL")
    
    if polling_mode or not webhook_url:
        logger.info("Starting ExhibitLedger bot in POLLING mode")
        app_bot = build_application()
        app_bot.run_polling(
            allowed_updates=Update.ALL_TYPES,
            drop_pending_updates=True,
        )
    else:
        logger.info("Starting ExhibitLedger bot in WEBHOOK mode")
        uvicorn.run(app, host="0.0.0.0", port=port)

if __name__ == "__main__":
    main()

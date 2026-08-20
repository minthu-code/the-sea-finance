import os
import re
import sqlite3
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_DB_PATH = os.environ.get("DB_PATH") or os.environ.get("DATABASE_PATH") or "./exhibitledger.db"

ACCOUNT_HEADS = [
    {
        "name": "Transport & Local Logistics",
        "section": "direct_cost",
        "keywords": ["transport", "taxi", "truck", "van", "grab", "delivery", "logistics", "local", "moving"],
    },
    {
        "name": "Air Cargo & Freight",
        "section": "operating_expense",
        "keywords": ["air cargo", "freight", "cargo", "shipping", "courier", "dhl", "fedex", "ems", "export", "import"],
    },
    {
        "name": "Customs Duties & Taxes",
        "section": "operating_expense",
        "keywords": ["customs", "duty", "duties", "tax", "vat", "import tax", "export tax", "clearance"],
    },
    {
        "name": "Venue Rental",
        "section": "operating_expense",
        "keywords": ["venue", "rental", "rent", "space", "hall", "booth", "gallery hire"],
    },
    {
        "name": "Installation & Production",
        "section": "direct_cost",
        "keywords": ["installation", "install", "production", "lighting", "plinth", "wall", "label", "catalog", "printing", "setup"],
    },
    {
        "name": "Framing & Artwork Preparation",
        "section": "direct_cost",
        "keywords": ["frame", "framing", "canvas", "stretch", "varnish", "mount", "artwork preparation", "preparation"],
    },
    {
        "name": "Repairs & Conservation",
        "section": "direct_cost",
        "keywords": ["repair", "restore", "restoration", "conservation", "damage", "touch up", "fix"],
    },
    {
        "name": "Travel & Accommodation",
        "section": "operating_expense",
        "keywords": ["flight", "air ticket", "ticket", "hotel", "accommodation", "travel", "visa", "per diem", "taxi airport"],
    },
    {
        "name": "Insurance",
        "section": "operating_expense",
        "keywords": ["insurance", "insured", "policy", "premium", "coverage"],
    },
    {
        "name": "Security",
        "section": "operating_expense",
        "keywords": ["security", "guard", "cctv", "safety", "supervision"],
    },
    {
        "name": "Food & Beverage / Hospitality",
        "section": "operating_expense",
        "keywords": ["food", "drink", "coffee", "snack", "catering", "wine", "water", "hospitality", "beverage", "meal"],
    },
    {
        "name": "Opening Event / VIP Relations",
        "section": "operating_expense",
        "keywords": ["opening", "vip", "guest", "invitation", "ceremony", "reception", "flowers", "gift"],
    },
    {
        "name": "Marketing & PR",
        "section": "operating_expense",
        "keywords": ["marketing", "pr", "advert", "ads", "facebook", "instagram", "boost", "poster", "media", "press", "photographer", "photo"],
    },
    {
        "name": "Office & Admin Supplies",
        "section": "operating_expense",
        "keywords": ["office", "admin", "paper", "ink", "stationery", "supplies", "receipt book", "folder"],
    },
    {
        "name": "Staff & Helpers / Labor",
        "section": "operating_expense",
        "keywords": ["staff", "helper", "labor", "labour", "assistant", "wage", "salary", "overtime", "crew"],
    },
    {
        "name": "Banking & Payment Fees",
        "section": "operating_expense",
        "keywords": ["bank", "fee", "fees", "transfer", "payment", "credit card", "card", "promptpay", "charge"],
    },
    {
        "name": "Miscellaneous (Needs Review)",
        "section": "operating_expense",
        "keywords": [],
    },
]
ACCOUNT_HEAD_BY_NAME = {row["name"].lower(): row for row in ACCOUNT_HEADS}
ALLOWED_PARTY_TYPES = {"gallery", "artist", "collaborator", "collector"}


def db_path() -> str:
    return os.environ.get("DB_PATH") or os.environ.get("DATABASE_PATH") or DEFAULT_DB_PATH


@contextmanager
def connect(path: str | None = None):
    conn = sqlite3.connect(path or db_path())
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def _table_columns(conn: sqlite3.Connection, table_name: str) -> set[str]:
    return {row[1] for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()}


def _utc_now() -> str:
    return datetime.utcnow().isoformat(timespec="seconds") + "Z"


def _insert_audit(conn: sqlite3.Connection, action: str, exhibition_code: str | None, details: str) -> None:
    conn.execute(
        "INSERT INTO audit_log (timestamp, action, exhibition_code, details) VALUES (?, ?, ?, ?)",
        (_utc_now(), action, exhibition_code, details),
    )


def init_db(path: str | None = None) -> None:
    """Create all legacy and workflow tables without destroying existing data."""
    with connect(path) as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS exhibitions (
                code TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                location TEXT,
                start_date TEXT,
                end_date TEXT,
                status TEXT DEFAULT 'active',
                currency TEXT DEFAULT 'THB',
                notes TEXT
            );

            CREATE TABLE IF NOT EXISTS pnl_lines (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                exhibition_code TEXT NOT NULL,
                section TEXT NOT NULL,
                category TEXT NOT NULL,
                description TEXT,
                amount_thb REAL NOT NULL DEFAULT 0,
                source_amount REAL,
                source_currency TEXT,
                source_ref TEXT,
                sort_order INTEGER DEFAULT 0,
                FOREIGN KEY (exhibition_code) REFERENCES exhibitions(code)
            );

            CREATE TABLE IF NOT EXISTS artist_payables (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                exhibition_code TEXT NOT NULL,
                artist TEXT NOT NULL,
                invoice_ref TEXT,
                gross_sale_thb REAL NOT NULL DEFAULT 0,
                gallery_commission_thb REAL NOT NULL DEFAULT 0,
                artist_payable_thb REAL NOT NULL DEFAULT 0,
                paid_thb REAL NOT NULL DEFAULT 0,
                outstanding_thb REAL NOT NULL DEFAULT 0,
                status TEXT DEFAULT 'Pending',
                source_amount REAL,
                source_currency TEXT,
                notes TEXT,
                FOREIGN KEY (exhibition_code) REFERENCES exhibitions(code)
            );

            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT NOT NULL,
                action TEXT NOT NULL,
                exhibition_code TEXT,
                details TEXT
            );

            CREATE TABLE IF NOT EXISTS commission_split_rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                exhibition_code TEXT NOT NULL,
                party_type TEXT NOT NULL,
                party_name TEXT NOT NULL,
                percent REAL NOT NULL,
                sort_order INTEGER DEFAULT 0,
                active INTEGER DEFAULT 1,
                created_at TEXT NOT NULL,
                FOREIGN KEY (exhibition_code) REFERENCES exhibitions(code)
            );

            CREATE TABLE IF NOT EXISTS artworks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                exhibition_code TEXT NOT NULL,
                title TEXT NOT NULL,
                artist TEXT NOT NULL,
                asking_price_thb REAL NOT NULL DEFAULT 0,
                status TEXT DEFAULT 'available',
                created_at TEXT NOT NULL,
                sold_at TEXT,
                FOREIGN KEY (exhibition_code) REFERENCES exhibitions(code)
            );

            CREATE TABLE IF NOT EXISTS artwork_sales (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                artwork_id INTEGER NOT NULL,
                exhibition_code TEXT NOT NULL,
                actual_price_thb REAL NOT NULL,
                sale_date TEXT NOT NULL,
                gallery_share_thb REAL NOT NULL DEFAULT 0,
                collaborator_share_thb REAL NOT NULL DEFAULT 0,
                artist_payable_thb REAL NOT NULL DEFAULT 0,
                split_summary TEXT,
                buyer_name TEXT,
                amount_collected_thb REAL NOT NULL DEFAULT 0,
                payment_status TEXT DEFAULT 'pending', -- pending, partial, paid
                expected_payment_date TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (artwork_id) REFERENCES artworks(id),
                FOREIGN KEY (exhibition_code) REFERENCES exhibitions(code)
            );

            CREATE TABLE IF NOT EXISTS sale_allocations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sale_id INTEGER NOT NULL,
                exhibition_code TEXT NOT NULL,
                artwork_id INTEGER NOT NULL,
                party_type TEXT NOT NULL,
                party_name TEXT NOT NULL,
                percent REAL NOT NULL,
                amount_thb REAL NOT NULL,
                FOREIGN KEY (sale_id) REFERENCES artwork_sales(id),
                FOREIGN KEY (artwork_id) REFERENCES artworks(id),
                FOREIGN KEY (exhibition_code) REFERENCES exhibitions(code)
            );

            CREATE TABLE IF NOT EXISTS pending_expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                exhibition_code TEXT NOT NULL,
                raw_text TEXT,
                description TEXT,
                suggested_amount_thb REAL NOT NULL DEFAULT 0,
                suggested_account_head TEXT NOT NULL,
                suggested_section TEXT NOT NULL,
                status TEXT DEFAULT 'pending',
                photo_file_id TEXT,
                created_at TEXT NOT NULL,
                confirmed_at TEXT,
                FOREIGN KEY (exhibition_code) REFERENCES exhibitions(code)
            );

            CREATE TABLE IF NOT EXISTS confirmed_expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                exhibition_code TEXT NOT NULL,
                account_head TEXT NOT NULL,
                section TEXT NOT NULL,
                description TEXT,
                amount_thb REAL NOT NULL,
                receipt_ref TEXT,
                raw_text TEXT,
                artist_tag TEXT, -- Link expense to a specific artist for ROI
                pending_expense_id INTEGER,
                created_at TEXT NOT NULL,
                pnl_line_id INTEGER,
                FOREIGN KEY (exhibition_code) REFERENCES exhibitions(code),
                FOREIGN KEY (pending_expense_id) REFERENCES pending_expenses(id),
                FOREIGN KEY (pnl_line_id) REFERENCES pnl_lines(id)
            );


            CREATE TABLE IF NOT EXISTS expense_budgets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                exhibition_code TEXT NOT NULL,
                account_head TEXT NOT NULL,
                section TEXT NOT NULL,
                budget_thb REAL NOT NULL DEFAULT 0,
                notes TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(exhibition_code, account_head),
                FOREIGN KEY (exhibition_code) REFERENCES exhibitions(code)
            );

            CREATE TABLE IF NOT EXISTS user_states (
                chat_id INTEGER PRIMARY KEY,
                current_exhibition TEXT,
                active_flow TEXT,
                flow_step INTEGER DEFAULT 0,
                flow_data TEXT DEFAULT '{}',
                updated_at TEXT
            );
            """
        )

        sale_column_additions = {
            "buyer_name": "TEXT",
            "payment_status": "TEXT DEFAULT 'collected'",
            "amount_collected_thb": "REAL DEFAULT 0",
            "balance_due_thb": "REAL DEFAULT 0",
            "payment_method": "TEXT",
            "notes": "TEXT",
        }
        sale_columns = _table_columns(conn, "artwork_sales")
        for column, definition in sale_column_additions.items():
            if column not in sale_columns:
                conn.execute(f"ALTER TABLE artwork_sales ADD COLUMN {column} {definition}")

        artwork_column_additions = {
            "inventory_code": "TEXT",
            "medium": "TEXT",
            "dimensions": "TEXT",
            "year_created": "TEXT",
        }
        artwork_columns = _table_columns(conn, "artworks")
        for column, definition in artwork_column_additions.items():
            if column not in artwork_columns:
                conn.execute(f"ALTER TABLE artworks ADD COLUMN {column} {definition}")

        # Older databases were created before expenses could be linked to an artist.
        # Add the column safely so receipt logging and Artist ROI work on existing installs.
        pending_expense_columns = _table_columns(conn, "pending_expenses")
        if "artist_tag" not in pending_expense_columns:
            conn.execute("ALTER TABLE pending_expenses ADD COLUMN artist_tag TEXT")
        for column in ("recipient", "category"):
            if column not in pending_expense_columns:
                conn.execute(f"ALTER TABLE pending_expenses ADD COLUMN {column} TEXT")

        # Recipient (vendor paid) and Category (sub-classification under an account head) —
        # needed for a proper accounting-firm-style ledger export.
        confirmed_expense_columns = _table_columns(conn, "confirmed_expenses")
        for column in ("recipient", "category"):
            if column not in confirmed_expense_columns:
                conn.execute(f"ALTER TABLE confirmed_expenses ADD COLUMN {column} TEXT")



def money(amount: float) -> str:
    return f"฿{float(amount or 0):,.2f} THB"


def compact_money(amount: float) -> str:
    return f"฿{float(amount or 0):,.0f}"


def normalize_code(code: str) -> str:
    cleaned = (code or "").strip().upper()
    if not cleaned:
        raise ValueError("Exhibition code is required.")
    return cleaned


def list_exhibitions() -> List[sqlite3.Row]:
    with connect() as conn:
        return conn.execute(
            "SELECT code, name, location, start_date, end_date, status, currency FROM exhibitions ORDER BY start_date DESC, code"
        ).fetchall()


def get_exhibition(code: str) -> sqlite3.Row | None:
    with connect() as conn:
        return conn.execute("SELECT * FROM exhibitions WHERE UPPER(code) = UPPER(?)", (code,)).fetchone()


def create_exhibition(
    code: str,
    name: str,
    location: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    notes: str | None = None,
) -> sqlite3.Row:
    code = normalize_code(code)
    name = (name or "").strip()
    if not name:
        raise ValueError("Exhibition name is required.")
    with connect() as conn:
        existing = conn.execute("SELECT code FROM exhibitions WHERE UPPER(code) = UPPER(?)", (code,)).fetchone()
        if existing:
            raise ValueError(f"Exhibition already exists: {code}")
        now = _utc_now()
        conn.execute(
            """
            INSERT INTO exhibitions (code, name, location, start_date, end_date, status, currency, notes)
            VALUES (?, ?, ?, ?, ?, 'active', 'THB', ?)
            """,
            (code, name, location, start_date, end_date, notes),
        )
        # Every exhibition starts with an explicit, editable 50/50 gallery/artist rule
        # so recording a sale never fails silently because the setup is incomplete.
        conn.executemany(
            """
            INSERT INTO commission_split_rules
            (exhibition_code, party_type, party_name, percent, sort_order, active, created_at)
            VALUES (?, ?, ?, ?, ?, 1, ?)
            """,
            [
                (code, "gallery", "Gallery", 50.0, 1, now),
                (code, "artist", "Artist", 50.0, 2, now),
            ],
        )
        _insert_audit(conn, "create_exhibition", code, f"Created exhibition {name} with default 50/50 commission split")
        return conn.execute("SELECT * FROM exhibitions WHERE code = ?", (code,)).fetchone()


def get_lines(code: str) -> List[sqlite3.Row]:
    with connect() as conn:
        return conn.execute(
            """
            SELECT * FROM pnl_lines
            WHERE UPPER(exhibition_code) = UPPER(?)
            ORDER BY sort_order, id
            """,
            (code,),
        ).fetchall()


def get_artist_payables(code: str) -> List[sqlite3.Row]:
    with connect() as conn:
        return conn.execute(
            """
            SELECT * FROM artist_payables
            WHERE UPPER(exhibition_code) = UPPER(?)
            ORDER BY outstanding_thb DESC, artist
            """,
            (code,),
        ).fetchall()


def section_total(lines: List[sqlite3.Row], section: str) -> float:
    return sum(float(row["amount_thb"] or 0) for row in lines if row["section"] == section)


def calculate_report(code: str) -> Dict:
    exhibition = get_exhibition(code)
    if not exhibition:
        raise ValueError(f"Exhibition not found: {code}")

    lines = get_lines(code)
    payables = get_artist_payables(code)

    gross_sales = section_total(lines, "sales_bridge")
    gallery_revenue = section_total(lines, "gallery_revenue")
    direct_costs = section_total(lines, "direct_cost")
    operating_expenses = section_total(lines, "operating_expense")
    overhead = section_total(lines, "allocated_overhead")
    gross_profit = gallery_revenue - direct_costs
    contribution_profit = gross_profit - operating_expenses
    net_profit = contribution_profit - overhead

    artist_payable_total = sum(float(row["artist_payable_thb"] or 0) for row in payables)
    artist_paid_total = sum(float(row["paid_thb"] or 0) for row in payables)
    artist_outstanding_total = sum(float(row["outstanding_thb"] or 0) for row in payables)

    gross_margin_pct = (gross_profit / gallery_revenue * 100) if gallery_revenue else 0
    contribution_margin_pct = (contribution_profit / gallery_revenue * 100) if gallery_revenue else 0
    net_margin_pct = (net_profit / gallery_revenue * 100) if gallery_revenue else 0
    expense_ratio_pct = ((direct_costs + operating_expenses + overhead) / gallery_revenue * 100) if gallery_revenue else 0

    return {
        "exhibition": dict(exhibition),
        "lines": [dict(row) for row in lines],
        "payables": [dict(row) for row in payables],
        "totals": {
            "gross_sales": gross_sales,
            "gallery_revenue": gallery_revenue,
            "direct_costs": direct_costs,
            "operating_expenses": operating_expenses,
            "allocated_overhead": overhead,
            "gross_profit": gross_profit,
            "contribution_profit": contribution_profit,
            "net_profit": net_profit,
            "artist_payable_total": artist_payable_total,
            "artist_paid_total": artist_paid_total,
            "artist_outstanding_total": artist_outstanding_total,
            "gross_margin_pct": gross_margin_pct,
            "contribution_margin_pct": contribution_margin_pct,
            "net_margin_pct": net_margin_pct,
            "expense_ratio_pct": expense_ratio_pct,
        },
    }


def grouped_lines(report: Dict, section: str) -> List[Dict]:
    return [row for row in report["lines"] if row["section"] == section]


def format_report_markdown(code: str) -> str:
    report = calculate_report(code)
    ex = report["exhibition"]
    totals = report["totals"]

    def line_items(section: str) -> str:
        rows = grouped_lines(report, section)
        if not rows:
            return "_No lines recorded._\n"
        return "\n".join(f"• {row['category']}: {money(float(row['amount_thb']))}" for row in rows) + "\n"

    generated = datetime.now().strftime("%Y-%m-%d %H:%M")
    text = [
        "*THE SEA ART GALLERY*",
        "*Profit & Loss Statement*",
        f"Exhibition: *{ex['name']}* (`{ex['code']}`)",
        f"Location: {ex['location'] or '-'}",
        f"Period: {ex['start_date'] or '-'} to {ex['end_date'] or '-'}",
        "Currency: *THB only*",
        f"Generated: {generated}",
        "",
        "*A. Sales Bridge*",
        line_items("sales_bridge"),
        f"Gross Artwork Sales / Activity: *{money(totals['gross_sales'])}*",
        "",
        "*B. Gallery Revenue*",
        line_items("gallery_revenue"),
        f"Total Gallery Revenue: *{money(totals['gallery_revenue'])}*",
        "",
        "*C. Direct Costs*",
        line_items("direct_cost"),
        f"Total Direct Costs: *{money(totals['direct_costs'])}*",
        "",
        f"*Gross Profit:* *{money(totals['gross_profit'])}* ({totals['gross_margin_pct']:.1f}%)",
        "",
        "*D. Exhibition Operating Expenses*",
        line_items("operating_expense"),
        f"Total Operating Expenses: *{money(totals['operating_expenses'])}*",
        "",
        f"*Contribution Profit:* *{money(totals['contribution_profit'])}* ({totals['contribution_margin_pct']:.1f}%)",
        "",
        "*E. Allocated Overhead*",
        line_items("allocated_overhead"),
        f"Allocated Overhead: *{money(totals['allocated_overhead'])}*",
        "",
        f"*NET PROFIT / (LOSS):* *{money(totals['net_profit'])}* ({totals['net_margin_pct']:.1f}%)",
        "",
        "*Artist Payable Control*",
        f"Artist Payable Total: *{money(totals['artist_payable_total'])}*",
        f"Paid: *{money(totals['artist_paid_total'])}*",
        f"Outstanding: *{money(totals['artist_outstanding_total'])}*",
        "",
        "_Note: This report is exhibition-by-exhibition and THB-only. Confirm receipt classifications and commission splits before relying on final numbers._",
    ]
    return "\n".join(text)


def format_artist_payables_markdown(code: str) -> str:
    exhibition = get_exhibition(code)
    if not exhibition:
        raise ValueError(f"Exhibition not found: {code}")
    payables = get_artist_payables(code)
    if not payables:
        return f"No artist payable rows found for `{code}`."

    lines = [f"*Artist Payables — {exhibition['name']}*", ""]
    for row in payables:
        lines.append(
            f"• *{row['artist']}* — Gross {compact_money(float(row['gross_sale_thb']))}; "
            f"Gallery commission {compact_money(float(row['gallery_commission_thb']))}; "
            f"Artist payable {compact_money(float(row['artist_payable_thb']))}; "
            f"Outstanding {compact_money(float(row['outstanding_thb']))}; Status: {row['status']}"
        )
    return "\n".join(lines)


def data_quality_checks(code: str) -> List[str]:
    report = calculate_report(code)
    warnings: List[str] = []
    ex = report["exhibition"]
    lines = report["lines"]
    payables = report["payables"]
    totals = report["totals"]

    if ex.get("currency") != "THB":
        warnings.append("Exhibition currency is not THB.")
    if not lines:
        warnings.append("No P&L lines recorded.")
    if not any(row["section"] == "gallery_revenue" for row in lines):
        warnings.append("No gallery revenue lines recorded.")
    if not any(row["section"] == "sales_bridge" for row in lines):
        warnings.append("No gross sales bridge line recorded.")
    if any((row.get("source_currency") or "THB") != "THB" for row in lines):
        warnings.append("Some rows were imported from non-THB sources. Confirm conversion rate before final use.")

    if payables:
        payable_outstanding_sum = sum(float(r["outstanding_thb"] or 0) for r in payables)
        payable_gross_sum = sum(float(r["gross_sale_thb"] or 0) for r in payables)
        payable_commission_sum = sum(float(r["gallery_commission_thb"] or 0) for r in payables)
        payable_artist_sum = sum(float(r["artist_payable_thb"] or 0) for r in payables)
        if abs(totals["artist_outstanding_total"] - payable_outstanding_sum) > 0.01:
            warnings.append("Artist payable outstanding total does not reconcile.")
        if totals["gross_sales"] and abs(totals["gross_sales"] - payable_gross_sum) > max(1.0, totals["gross_sales"] * 0.02):
            warnings.append("Artist commission records appear partial or unmapped; gross sales bridge does not yet reconcile to all artist sale records.")
        split_rows_for_reconcile = get_split_rules(code)
        collaborator_pct = sum(float(r["percent"] or 0) for r in split_rows_for_reconcile if str(r["party_type"]).lower() not in {"gallery", "artist"})
        expected_gallery_artist_share = payable_gross_sum * max(0.0, (100.0 - collaborator_pct)) / 100.0
        if abs((payable_commission_sum + payable_artist_sum) - expected_gallery_artist_share) > max(1.0, payable_gross_sum * 0.02):
            warnings.append("Gallery and artist portions do not reconcile to the active split rule.")

    split_rows = get_split_rules(code)
    if not split_rows:
        warnings.append("No active commission split rule has been set for this exhibition.")
    else:
        split_total = sum(float(r["percent"] or 0) for r in split_rows)
        if abs(split_total - 100) > 0.01:
            warnings.append(f"Commission split totals {split_total:.2f}%, not 100%.")

    pending_count = count_pending_expenses(code)
    if pending_count:
        warnings.append(f"There are {pending_count} pending expense receipt(s) awaiting approval.")

    metrics = calculate_inventory_metrics(code)
    if metrics["total_artworks"] == 0:
        warnings.append("No artwork inventory has been registered for this exhibition.")
    if metrics["receivables_thb"] > 0:
        warnings.append(f"There are uncollected or partially collected sale balances totaling {money(metrics['receivables_thb'])}.")
    if metrics["sold_artworks"] > 0 and metrics["total_artworks"] > 0 and metrics["sell_through_rate_pct"] < 30:
        warnings.append(f"Sell-through is currently {metrics['sell_through_rate_pct']:.1f}%; review whether final reporting should mention unsold inventory.")

    budget_rows = calculate_budget_report(code)
    over_budget = [row for row in budget_rows if row["budget_thb"] > 0 and row["variance_thb"] < 0]
    if over_budget:
        worst = sorted(over_budget, key=lambda row: row["variance_thb"])[0]
        warnings.append(f"Budget overrun: {worst['account_head']} is over budget by {money(abs(worst['variance_thb']))}.")

    if totals["gallery_revenue"] <= 0:
        warnings.append("Gallery revenue is zero or negative.")
    if totals["direct_costs"] < 0 or totals["operating_expenses"] < 0 or totals["allocated_overhead"] < 0:
        warnings.append("One or more cost sections are negative; confirm sign convention.")
    if totals["net_profit"] < 0:
        warnings.append("Exhibition currently reports a net loss under current assumptions.")
    if totals["expense_ratio_pct"] > 100:
        warnings.append("Total expense ratio is above 100% of gallery revenue; review cost classification.")
    if not warnings:
        warnings.append("No blocking issues found for current exhibition data.")
    return warnings


def format_executive_summary_markdown(code: str) -> str:
    report = calculate_report(code)
    ex = report["exhibition"]
    totals = report["totals"]
    metrics = calculate_inventory_metrics(code)
    checks = data_quality_checks(code)
    status = "PROFIT" if totals["net_profit"] >= 0 else "LOSS"
    text = [
        f"*Executive Dashboard — {ex['name']}*",
        f"Code: `{ex['code']}`",
        "Currency: *THB only*",
        "",
        f"P&L status: *{status}*",
        f"Gross Artwork Sales / Activity: *{money(totals['gross_sales'])}*",
        f"Gallery Revenue: *{money(totals['gallery_revenue'])}*",
        f"Direct Costs: *{money(totals['direct_costs'])}*",
        f"Operating Expenses: *{money(totals['operating_expenses'])}*",
        f"Net Profit / (Loss): *{money(totals['net_profit'])}*",
        f"Net Margin on Gallery Revenue: *{totals['net_margin_pct']:.1f}%*",
        "",
        "Inventory and collection controls:",
        f"• Artworks registered: {metrics['total_artworks']} | Sold: {metrics['sold_artworks']} | Available: {metrics['available_artworks']}",
        f"• Sell-through rate: {metrics['sell_through_rate_pct']:.1f}%",
        f"• Unsold asking value: {money(metrics['unsold_asking_value_thb'])}",
        f"• Cash collected from sales: {money(metrics['cash_collected_thb'])}",
        f"• Sale receivables outstanding: {money(metrics['receivables_thb'])}",
        f"• Pending receipts awaiting approval: {metrics['pending_receipts']}",
        "",
        "Control points:",
    ]
    text.extend(f"• {check}" for check in checks[:8])
    text.append("")
    text.append("Recommended next action: clear pending receipts, collect outstanding sale balances, verify account heads, and run /export before final review.")
    return "\n".join(text)


def log_action(action: str, exhibition_code: str | None, details: str) -> None:
    with connect() as conn:
        _insert_audit(conn, action, exhibition_code, details)


# ---------------------------------------------------------------------------
# Commission split, artwork, and sale workflow
# ---------------------------------------------------------------------------


def set_commission_splits(exhibition_code: str, entries: Sequence[Dict]) -> List[sqlite3.Row]:
    exhibition_code = normalize_code(exhibition_code)
    if not get_exhibition(exhibition_code):
        raise ValueError(f"Exhibition not found: {exhibition_code}")
    if not entries:
        raise ValueError("At least one split entry is required.")

    cleaned = []
    for idx, entry in enumerate(entries, start=1):
        party_type = (entry.get("party_type") or "").strip().lower()
        party_name = (entry.get("party_name") or party_type.title()).strip()
        try:
            percent = float(entry.get("percent"))
        except (TypeError, ValueError):
            raise ValueError(f"Invalid percent for split entry {entry!r}")
        if party_type not in ALLOWED_PARTY_TYPES:
            raise ValueError(f"Unsupported party type: {party_type}. Use gallery, artist, collaborator, or collector.")
        if percent <= 0:
            raise ValueError("Split percentages must be greater than zero.")
        cleaned.append({"party_type": party_type, "party_name": party_name, "percent": percent, "sort_order": idx})

    total = sum(row["percent"] for row in cleaned)
    if abs(total - 100.0) > 0.01:
        raise ValueError(f"Split percentages must total 100%. Current total is {total:.2f}%.")

    with connect() as conn:
        conn.execute("DELETE FROM commission_split_rules WHERE UPPER(exhibition_code) = UPPER(?)", (exhibition_code,))
        for row in cleaned:
            conn.execute(
                """
                INSERT INTO commission_split_rules
                (exhibition_code, party_type, party_name, percent, sort_order, active, created_at)
                VALUES (?, ?, ?, ?, ?, 1, ?)
                """,
                (exhibition_code, row["party_type"], row["party_name"], row["percent"], row["sort_order"], _utc_now()),
            )
        _insert_audit(conn, "set_commission_splits", exhibition_code, format_split_summary(cleaned))
        return conn.execute(
            "SELECT * FROM commission_split_rules WHERE UPPER(exhibition_code) = UPPER(?) ORDER BY sort_order, id",
            (exhibition_code,),
        ).fetchall()


def get_split_rules(exhibition_code: str) -> List[sqlite3.Row]:
    with connect() as conn:
        return conn.execute(
            """
            SELECT * FROM commission_split_rules
            WHERE UPPER(exhibition_code) = UPPER(?) AND active = 1
            ORDER BY sort_order, id
            """,
            (exhibition_code,),
        ).fetchall()


def format_split_summary(entries: Sequence[Dict | sqlite3.Row]) -> str:
    parts = []
    for row in entries:
        party_type = row["party_type"]
        party_name = row["party_name"]
        percent = float(row["percent"])
        parts.append(f"{party_type}:{party_name} {percent:g}%")
    return "; ".join(parts)


def format_split_rules_markdown(exhibition_code: str) -> str:
    exhibition = get_exhibition(exhibition_code)
    if not exhibition:
        raise ValueError(f"Exhibition not found: {exhibition_code}")
    rows = get_split_rules(exhibition_code)
    if not rows:
        return f"No commission split rule has been set for `{normalize_code(exhibition_code)}`."
    total = sum(float(row["percent"] or 0) for row in rows)
    lines = [f"*Commission Split — {exhibition['name']}*", ""]
    for row in rows:
        lines.append(f"• {row['party_type'].title()} — {row['party_name']}: {float(row['percent']):g}%")
    lines.append("")
    lines.append(f"Total: *{total:g}%*")
    return "\n".join(lines)


def add_artwork(exhibition_code: str, title: str, artist: str, asking_price_thb: float) -> sqlite3.Row:
    exhibition_code = normalize_code(exhibition_code)
    if not get_exhibition(exhibition_code):
        raise ValueError(f"Exhibition not found: {exhibition_code}")
    title = (title or "").strip()
    artist = (artist or "").strip()
    price = float(asking_price_thb)
    if not title or not artist:
        raise ValueError("Artwork title and artist are required.")
    if price < 0:
        raise ValueError("Artwork price cannot be negative.")
    with connect() as conn:
        cur = conn.execute(
            """
            INSERT INTO artworks (exhibition_code, title, artist, asking_price_thb, status, created_at)
            VALUES (?, ?, ?, ?, 'available', ?)
            """,
            (exhibition_code, title, artist, price, _utc_now()),
        )
        artwork_id = cur.lastrowid
        _insert_audit(conn, "add_artwork", exhibition_code, f"Added artwork #{artwork_id}: {title} by {artist} at {money(price)}")
        return conn.execute("SELECT * FROM artworks WHERE id = ?", (artwork_id,)).fetchone()


def bulk_add_artworks(exhibition_code: str, artworks: List[Dict]) -> List[int]:
    """Import multiple artworks at once."""
    exhibition_code = normalize_code(exhibition_code)
    if not get_exhibition(exhibition_code):
        raise ValueError(f"Exhibition not found: {exhibition_code}")
    ids = []
    with connect() as conn:
        for art in artworks:
            title = (art.get("title") or "Untitled").strip()
            artist = (art.get("artist") or "Unknown").strip()
            price = float(art.get("price") or 0)
            
            cur = conn.execute(
                """
                INSERT INTO artworks (exhibition_code, title, artist, asking_price_thb, status, created_at)
                VALUES (?, ?, ?, ?, 'available', ?)
                """,
                (exhibition_code, title, artist, price, _utc_now()),
            )
            ids.append(cur.lastrowid)
        
        _insert_audit(conn, "bulk_add_artworks", exhibition_code, f"Imported {len(artworks)} artworks")
    return ids


def get_artwork(artwork_id: int) -> sqlite3.Row | None:
    with connect() as conn:
        return conn.execute("SELECT * FROM artworks WHERE id = ?", (int(artwork_id),)).fetchone()


def list_artworks(exhibition_code: str, include_sold: bool = True) -> List[sqlite3.Row]:
    with connect() as conn:
        if include_sold:
            return conn.execute(
                """
                SELECT * FROM artworks WHERE UPPER(exhibition_code) = UPPER(?)
                ORDER BY id
                """,
                (exhibition_code,),
            ).fetchall()
        return conn.execute(
            """
            SELECT * FROM artworks WHERE UPPER(exhibition_code) = UPPER(?) AND status <> 'sold'
            ORDER BY id
            """,
            (exhibition_code,),
        ).fetchall()


def format_artworks_markdown(exhibition_code: str) -> str:
    exhibition = get_exhibition(exhibition_code)
    if not exhibition:
        raise ValueError(f"Exhibition not found: {exhibition_code}")
    rows = list_artworks(exhibition_code)
    if not rows:
        return f"No artworks have been registered for `{normalize_code(exhibition_code)}`."
    lines = [f"*Artworks — {exhibition['name']}*", ""]
    for row in rows:
        lines.append(
            f"• #{row['id']} — {row['title']} / {row['artist']} / Asking {compact_money(row['asking_price_thb'])} / {row['status']}"
        )
    return "\n".join(lines)


def _allocation_amounts(actual_price_thb: float, split_rows: Sequence[sqlite3.Row]) -> List[Dict]:
    allocations = []
    running = 0.0
    for idx, row in enumerate(split_rows):
        if idx == len(split_rows) - 1:
            amount = round(actual_price_thb - running, 2)
        else:
            amount = round(actual_price_thb * float(row["percent"]) / 100.0, 2)
            running += amount
        allocations.append(
            {
                "party_type": row["party_type"],
                "party_name": row["party_name"],
                "percent": float(row["percent"]),
                "amount_thb": amount,
            }
        )
    return allocations


def record_sale(
    artwork_id: int,
    actual_price_thb: float,
    sale_date: str | None = None,
    buyer_name: str | None = None,
    amount_collected_thb: float | None = None,
    payment_method: str | None = None,
    notes: str | None = None,
) -> Dict:
    artwork_id = int(artwork_id)
    actual_price = float(actual_price_thb)
    if actual_price <= 0:
        raise ValueError("Actual sale price must be greater than zero.")
    artwork = get_artwork(artwork_id)
    if not artwork:
        raise ValueError(f"Artwork not found: {artwork_id}")
    if artwork["status"] == "sold":
        raise ValueError(f"Artwork #{artwork_id} is already marked as sold.")
    exhibition_code = artwork["exhibition_code"]
    splits = get_split_rules(exhibition_code)
    if not splits:
        raise ValueError(f"Set a commission split first with /set_split {exhibition_code} ...")

    sale_date = sale_date or datetime.now().strftime("%Y-%m-%d")
    collected = actual_price if amount_collected_thb is None else float(amount_collected_thb)
    if collected < 0:
        raise ValueError("Collected amount cannot be negative.")
    if collected > actual_price:
        raise ValueError("Collected amount cannot exceed actual sale price.")
    balance_due = round(actual_price - collected, 2)
    if balance_due <= 0:
        payment_status = "collected"
    elif collected > 0:
        payment_status = "partial"
    else:
        payment_status = "uncollected"

    allocations = _allocation_amounts(actual_price, splits)
    gallery_share = sum(row["amount_thb"] for row in allocations if row["party_type"] == "gallery")
    collaborator_share = sum(row["amount_thb"] for row in allocations if row["party_type"] in {"collaborator", "collector"})
    artist_payable = sum(row["amount_thb"] for row in allocations if row["party_type"] == "artist")
    split_summary = format_split_summary(allocations)

    with connect() as conn:
        cur = conn.execute(
            """
            INSERT INTO artwork_sales
            (artwork_id, exhibition_code, actual_price_thb, sale_date, gallery_share_thb, collaborator_share_thb,
             artist_payable_thb, split_summary, created_at, buyer_name, payment_status, amount_collected_thb,
             balance_due_thb, payment_method, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                artwork_id,
                exhibition_code,
                actual_price,
                sale_date,
                gallery_share,
                collaborator_share,
                artist_payable,
                split_summary,
                _utc_now(),
                (buyer_name or "").strip() or None,
                payment_status,
                collected,
                balance_due,
                (payment_method or "").strip() or None,
                (notes or "").strip() or None,
            ),
        )
        sale_id = cur.lastrowid
        for row in allocations:
            party_name = row["party_name"]
            if row["party_type"] == "artist" and party_name.lower() in {"artist", "default artist"}:
                party_name = artwork["artist"]
            conn.execute(
                """
                INSERT INTO sale_allocations
                (sale_id, exhibition_code, artwork_id, party_type, party_name, percent, amount_thb)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (sale_id, exhibition_code, artwork_id, row["party_type"], party_name, row["percent"], row["amount_thb"]),
            )

        conn.execute("UPDATE artworks SET status = 'sold', sold_at = ? WHERE id = ?", (_utc_now(), artwork_id))
        source_ref = f"sale:{sale_id}; artwork:{artwork_id}"
        conn.execute(
            """
            INSERT INTO pnl_lines (exhibition_code, section, category, description, amount_thb, source_amount, source_currency, source_ref, sort_order)
            VALUES (?, 'sales_bridge', 'Gross artwork sales', ?, ?, ?, 'THB', ?, 10)
            """,
            (exhibition_code, f"Sold artwork #{artwork_id}: {artwork['title']}", actual_price, actual_price, source_ref),
        )
        if gallery_share:
            conn.execute(
                """
                INSERT INTO pnl_lines (exhibition_code, section, category, description, amount_thb, source_amount, source_currency, source_ref, sort_order)
                VALUES (?, 'gallery_revenue', 'Gallery commission from sold artwork', ?, ?, ?, 'THB', ?, 20)
                """,
                (exhibition_code, f"Gallery share from artwork #{artwork_id}: {artwork['title']}", gallery_share, gallery_share, source_ref),
            )
        if artist_payable:
            conn.execute(
                """
                INSERT INTO pnl_lines (exhibition_code, section, category, description, amount_thb, source_amount, source_currency, source_ref, sort_order)
                VALUES (?, 'direct_cost', 'Artist payable from sold artwork', ?, ?, ?, 'THB', ?, 30)
                """,
                (exhibition_code, f"Artist share for {artwork['artist']} on artwork #{artwork_id}", artist_payable, artist_payable, source_ref),
            )
            conn.execute(
                """
                INSERT INTO artist_payables
                (exhibition_code, artist, invoice_ref, gross_sale_thb, gallery_commission_thb, artist_payable_thb,
                 paid_thb, outstanding_thb, status, source_amount, source_currency, notes)
                VALUES (?, ?, ?, ?, ?, ?, 0, ?, 'Pending', ?, 'THB', ?)
                """,
                (
                    exhibition_code,
                    artwork["artist"],
                    f"Sale #{sale_id} / Artwork #{artwork_id}",
                    actual_price,
                    gallery_share,
                    artist_payable,
                    artist_payable,
                    actual_price,
                    f"Auto-created from sale using split: {split_summary}; buyer={buyer_name or '-'}; payment={payment_status}",
                ),
            )
        if collaborator_share:
            conn.execute(
                """
                INSERT INTO pnl_lines (exhibition_code, section, category, description, amount_thb, source_amount, source_currency, source_ref, sort_order)
                VALUES (?, 'direct_cost', 'Collaborator / collector share from sold artwork', ?, ?, ?, 'THB', ?, 31)
                """,
                (exhibition_code, f"Collaborator or collector share from artwork #{artwork_id}", collaborator_share, collaborator_share, source_ref),
            )
        _insert_audit(conn, "record_sale", exhibition_code, f"Recorded sale #{sale_id} for artwork #{artwork_id} at {money(actual_price)}; collected {money(collected)}")
        sale = conn.execute("SELECT * FROM artwork_sales WHERE id = ?", (sale_id,)).fetchone()
        sale_allocations = conn.execute("SELECT * FROM sale_allocations WHERE sale_id = ? ORDER BY id", (sale_id,)).fetchall()
        return {"sale": dict(sale), "artwork": dict(artwork), "allocations": [dict(row) for row in sale_allocations]}


def format_sale_markdown(sale_result: Dict) -> str:
    sale = sale_result["sale"]
    artwork = sale_result["artwork"]
    lines = [
        f"*Sale Recorded — Artwork #{artwork['id']}*",
        f"Title: {artwork['title']}",
        f"Artist: {artwork['artist']}",
        f"Actual sale price: *{money(sale['actual_price_thb'])}*",
        f"Buyer: {sale.get('buyer_name') or '-'}",
        f"Payment status: *{sale.get('payment_status') or 'collected'}*",
        f"Collected: *{money(sale.get('amount_collected_thb') or sale['actual_price_thb'])}*",
        f"Balance due: *{money(sale.get('balance_due_thb') or 0)}*",
        "",
        "Split allocation:",
    ]
    for row in sale_result["allocations"]:
        lines.append(f"• {row['party_type'].title()} — {row['party_name']}: {float(row['percent']):g}% = {money(row['amount_thb'])}")
    lines.extend(
        [
            "",
            f"Gallery revenue posted: *{money(sale['gallery_share_thb'])}*",
            f"Artist payable posted: *{money(sale['artist_payable_thb'])}*",
            f"Collaborator / collector share posted: *{money(sale['collaborator_share_thb'])}*",
            "",
            "If payment status is partial or uncollected, the P&L still records the sale, while the dashboard shows the receivable balance.",
        ]
    )
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Receipt capture, approval, and expense reporting workflow
# ---------------------------------------------------------------------------


def parse_amount_thb(text: str | None) -> float:
    text = text or ""
    preferred = re.findall(r"(?:฿|THB|thb)\s*([0-9][0-9,]*(?:\.\d{1,2})?)", text)
    matches = preferred or re.findall(r"\b([0-9][0-9,]*(?:\.\d{1,2})?)\b", text)
    values = []
    for raw in matches:
        try:
            value = float(raw.replace(",", ""))
        except ValueError:
            continue
        if value > 0:
            values.append(value)
    if not values:
        return 0.0
    return round(max(values), 2)


def suggest_account_head(text: str | None) -> Dict:
    lowered = (text or "").lower()
    for row in ACCOUNT_HEADS:
        for keyword in row["keywords"]:
            if keyword in lowered:
                return row
    return ACCOUNT_HEADS[-1]


def account_head_names() -> List[str]:
    return [row["name"] for row in ACCOUNT_HEADS]


def get_account_head(name: str) -> Dict:
    row = ACCOUNT_HEAD_BY_NAME.get((name or "").strip().lower())
    if not row:
        raise ValueError(f"Unknown account head: {name}")
    return row


def clean_expense_description(raw_text: str | None) -> str:
    text = (raw_text or "Receipt photo / expense").strip()
    text = re.sub(r"(?:฿|THB|thb)?\s*[0-9][0-9,]*(?:\.\d{1,2})?", "", text).strip(" -:;,")
    return text[:180] or "Receipt photo / expense"


def create_pending_expense(
    exhibition_code: str,
    raw_text: str | None,
    photo_file_id: str | None = None,
    artist_tag: str | None = None,
    recipient: str | None = None,
    category: str | None = None,
) -> sqlite3.Row:
    exhibition_code = normalize_code(exhibition_code)
    if not get_exhibition(exhibition_code):
        raise ValueError(f"Exhibition not found: {exhibition_code}")
    amount = parse_amount_thb(raw_text)
    account = suggest_account_head(raw_text)
    description = clean_expense_description(raw_text)
    with connect() as conn:
        cur = conn.execute(
            """
            INSERT INTO pending_expenses
            (exhibition_code, raw_text, description, suggested_amount_thb, suggested_account_head, suggested_section,
             status, photo_file_id, artist_tag, recipient, category, created_at)
            VALUES (?, ?, ?, ?, ?, ?, 'pending', ?, ?, ?, ?, ?)
            """,
            (exhibition_code, raw_text, description, amount, account["name"], account["section"], photo_file_id, artist_tag,
             (recipient or "").strip() or None, (category or "").strip() or None, _utc_now()),
        )
        pending_id = cur.lastrowid
        _insert_audit(conn, "create_pending_expense", exhibition_code, f"Created pending expense #{pending_id}")
        return conn.execute("SELECT * FROM pending_expenses WHERE id = ?", (pending_id,)).fetchone()


def get_pending_expense(pending_id: int) -> sqlite3.Row | None:
    with connect() as conn:
        return conn.execute("SELECT * FROM pending_expenses WHERE id = ?", (int(pending_id),)).fetchone()


def list_pending_expenses(exhibition_code: str | None = None) -> List[sqlite3.Row]:
    with connect() as conn:
        if exhibition_code:
            return conn.execute(
                """
                SELECT * FROM pending_expenses
                WHERE UPPER(exhibition_code) = UPPER(?) AND status = 'pending'
                ORDER BY id
                """,
                (exhibition_code,),
            ).fetchall()
        return conn.execute("SELECT * FROM pending_expenses WHERE status = 'pending' ORDER BY exhibition_code, id").fetchall()


def count_pending_expenses(exhibition_code: str | None = None) -> int:
    return len(list_pending_expenses(exhibition_code))


def update_pending_account(pending_id: int, account_head: str) -> sqlite3.Row:
    account = get_account_head(account_head)
    with connect() as conn:
        row = conn.execute("SELECT * FROM pending_expenses WHERE id = ?", (int(pending_id),)).fetchone()
        if not row:
            raise ValueError(f"Pending expense not found: {pending_id}")
        if row["status"] != "pending":
            raise ValueError(f"Pending expense #{pending_id} is already {row['status']}.")
        conn.execute(
            "UPDATE pending_expenses SET suggested_account_head = ?, suggested_section = ? WHERE id = ?",
            (account["name"], account["section"], int(pending_id)),
        )
        _insert_audit(conn, "update_pending_account", row["exhibition_code"], f"Pending expense #{pending_id} account changed to {account['name']}")
        return conn.execute("SELECT * FROM pending_expenses WHERE id = ?", (int(pending_id),)).fetchone()


def update_pending_amount(pending_id: int, amount_thb: float) -> sqlite3.Row:
    amount = float(amount_thb)
    if amount <= 0:
        raise ValueError("Expense amount must be greater than zero.")
    with connect() as conn:
        row = conn.execute("SELECT * FROM pending_expenses WHERE id = ?", (int(pending_id),)).fetchone()
        if not row:
            raise ValueError(f"Pending expense not found: {pending_id}")
        if row["status"] != "pending":
            raise ValueError(f"Pending expense #{pending_id} is already {row['status']}.")
        conn.execute("UPDATE pending_expenses SET suggested_amount_thb = ? WHERE id = ?", (amount, int(pending_id)))
        _insert_audit(conn, "update_pending_amount", row["exhibition_code"], f"Pending expense #{pending_id} amount changed to {money(amount)}")
        return conn.execute("SELECT * FROM pending_expenses WHERE id = ?", (int(pending_id),)).fetchone()


def ignore_pending_expense(pending_id: int) -> sqlite3.Row:
    with connect() as conn:
        row = conn.execute("SELECT * FROM pending_expenses WHERE id = ?", (int(pending_id),)).fetchone()
        if not row:
            raise ValueError(f"Pending expense not found: {pending_id}")
        conn.execute("UPDATE pending_expenses SET status = 'ignored' WHERE id = ?", (int(pending_id),))
        _insert_audit(conn, "ignore_pending_expense", row["exhibition_code"], f"Ignored pending expense #{pending_id}")
        return conn.execute("SELECT * FROM pending_expenses WHERE id = ?", (int(pending_id),)).fetchone()


def confirm_pending_expense(pending_id: int) -> sqlite3.Row:
    with connect() as conn:
        pending = conn.execute("SELECT * FROM pending_expenses WHERE id = ?", (int(pending_id),)).fetchone()
        if not pending:
            raise ValueError(f"Pending expense not found: {pending_id}")
        if pending["status"] != "pending":
            raise ValueError(f"Pending expense #{pending_id} is already {pending['status']}.")
        amount = float(pending["suggested_amount_thb"] or 0)
        if amount <= 0:
            raise ValueError("Please change the amount before confirming. The current amount is zero.")
        source_ref = f"expense:{pending_id}"
        cur = conn.execute(
            """
            INSERT INTO pnl_lines
            (exhibition_code, section, category, description, amount_thb, source_amount, source_currency, source_ref, sort_order)
            VALUES (?, ?, ?, ?, ?, ?, 'THB', ?, 50)
            """,
            (
                pending["exhibition_code"],
                pending["suggested_section"],
                pending["suggested_account_head"],
                pending["description"],
                amount,
                amount,
                source_ref,
            ),
        )
        pnl_line_id = cur.lastrowid
        cur = conn.execute(
            """
            INSERT INTO confirmed_expenses
            (exhibition_code, account_head, section, description, amount_thb, receipt_ref, raw_text,
             artist_tag, recipient, category, pending_expense_id, created_at, pnl_line_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                pending["exhibition_code"],
                pending["suggested_account_head"],
                pending["suggested_section"],
                pending["description"],
                amount,
                source_ref,
                pending["raw_text"],
                pending["artist_tag"] if "artist_tag" in pending.keys() else None,
                pending["recipient"] if "recipient" in pending.keys() else None,
                pending["category"] if "category" in pending.keys() else None,
                int(pending_id),
                _utc_now(),
                pnl_line_id,
            ),
        )
        confirmed_id = cur.lastrowid
        conn.execute("UPDATE pending_expenses SET status = 'confirmed', confirmed_at = ? WHERE id = ?", (_utc_now(), int(pending_id)))
        _insert_audit(conn, "confirm_pending_expense", pending["exhibition_code"], f"Confirmed expense #{confirmed_id} from pending #{pending_id}")
        return conn.execute("SELECT * FROM confirmed_expenses WHERE id = ?", (confirmed_id,)).fetchone()


def list_confirmed_expenses(exhibition_code: str) -> List[sqlite3.Row]:
    with connect() as conn:
        return conn.execute(
            """
            SELECT * FROM confirmed_expenses
            WHERE UPPER(exhibition_code) = UPPER(?)
            ORDER BY account_head, id
            """,
            (exhibition_code,),
        ).fetchall()


def format_pending_expense_card(pending: sqlite3.Row) -> str:
    return "\n".join(
        [
            f"Receipt Pending Approval #{pending['id']}",
            f"Exhibition: {pending['exhibition_code']}",
            f"Amount: {money(pending['suggested_amount_thb'])}",
            f"Suggested Account Head: {pending['suggested_account_head']}",
            f"P&L Section: {pending['suggested_section']}",
            f"Description: {pending['description'] or '-'}",
            f"Status: {pending['status']}",
            "",
            "Please confirm, change the account head, change the amount, or ignore this receipt.",
        ]
    )


def format_pending_expenses_markdown(exhibition_code: str | None = None) -> str:
    rows = list_pending_expenses(exhibition_code)
    title = f"Pending Receipts — {normalize_code(exhibition_code)}" if exhibition_code else "Pending Receipts"
    if not rows:
        return f"{title}\n\nNo pending receipts awaiting approval."
    lines = [title, ""]
    for row in rows:
        lines.append(
            f"• #{row['id']} / {row['exhibition_code']} / {money(row['suggested_amount_thb'])} / "
            f"{row['suggested_account_head']} / {row['description']}"
        )
    return "\n".join(lines)


def format_expense_report_markdown(exhibition_code: str) -> str:
    exhibition = get_exhibition(exhibition_code)
    if not exhibition:
        raise ValueError(f"Exhibition not found: {exhibition_code}")
    rows = list_confirmed_expenses(exhibition_code)
    if not rows:
        return f"Expense Report — {exhibition['name']}\n\nNo confirmed expenses have been recorded."
    grouped: Dict[str, List[sqlite3.Row]] = {}
    for row in rows:
        grouped.setdefault(row["account_head"], []).append(row)
    grand_total = sum(float(row["amount_thb"] or 0) for row in rows)
    lines = [f"Expense Report — {exhibition['name']} ({exhibition['code']})", "Currency: THB only", ""]
    for account_head, account_rows in grouped.items():
        subtotal = sum(float(row["amount_thb"] or 0) for row in account_rows)
        lines.append(f"{account_head}: {money(subtotal)}")
        for row in account_rows:
            lines.append(f"  • #{row['id']} {row['description']} — {money(row['amount_thb'])}")
        lines.append("")
    lines.append(f"Total Confirmed Expenses: {money(grand_total)}")
    return "\n".join(lines)


def format_account_heads_markdown() -> str:
    lines = ["Expense Account Heads", "", "Use these classifications when approving receipts:", ""]
    for idx, row in enumerate(ACCOUNT_HEADS, start=1):
        lines.append(f"{idx}. {row['name']} — P&L section: {row['section']}")
    return "\n".join(lines)



# ---------------------------------------------------------------------------
# Management controls, budgets, inventory, and readiness dashboards
# ---------------------------------------------------------------------------


def list_sales(exhibition_code: str) -> List[sqlite3.Row]:
    with connect() as conn:
        return conn.execute(
            """
            SELECT s.*, a.title, a.artist, a.asking_price_thb
            FROM artwork_sales s
            LEFT JOIN artworks a ON a.id = s.artwork_id
            WHERE UPPER(s.exhibition_code) = UPPER(?)
            ORDER BY s.sale_date, s.id
            """,
            (exhibition_code,),
        ).fetchall()


def calculate_artist_roi(code: str) -> List[Dict]:
    """Calculate profitability per artist by deducting direct costs from gallery share."""
    with connect() as conn:
        # Get all artists in this exhibition
        artists = [row[0] for row in conn.execute("SELECT DISTINCT artist FROM artworks WHERE exhibition_code = ?", (code,)).fetchall()]
        
        roi_data = []
        for artist in artists:
            # Gallery share from sales of this artist's work
            gallery_share = conn.execute("""
                SELECT SUM(gallery_share_thb) FROM artwork_sales 
                JOIN artworks ON artwork_sales.artwork_id = artworks.id
                WHERE artworks.artist = ? AND artworks.exhibition_code = ?
            """, (artist, code)).fetchone()[0] or 0.0
            
            # Direct costs tagged to this artist
            direct_costs = conn.execute("""
                SELECT SUM(amount_thb) FROM confirmed_expenses
                WHERE artist_tag = ? AND exhibition_code = ?
            """, (artist, code)).fetchone()[0] or 0.0
            
            # Artworks stats
            total = conn.execute("SELECT COUNT(*) FROM artworks WHERE artist = ? AND exhibition_code = ?", (artist, code)).fetchone()[0]
            sold = conn.execute("SELECT COUNT(*) FROM artworks WHERE artist = ? AND exhibition_code = ? AND status = 'sold'", (artist, code)).fetchone()[0]
            
            roi_data.append({
                "artist": artist,
                "total_artworks": total,
                "sold_artworks": sold,
                "gallery_share": gallery_share,
                "direct_costs": direct_costs,
                "net_contribution": gallery_share - direct_costs,
                "sell_through_rate": (sold / total * 100) if total > 0 else 0
            })
            
        return sorted(roi_data, key=lambda x: x["net_contribution"], reverse=True)

def get_receivables_report(code: str) -> List[Dict]:
    """List all sales with outstanding balances and expected dates."""
    with connect() as conn:
        rows = conn.execute("""
            SELECT artwork_sales.*, artworks.title, artworks.artist 
            FROM artwork_sales
            JOIN artworks ON artwork_sales.artwork_id = artworks.id
            WHERE artwork_sales.exhibition_code = ? AND artwork_sales.payment_status != 'paid'
            ORDER BY artwork_sales.expected_payment_date ASC
        """, (code,)).fetchall()
        return [dict(r) for r in rows]

def get_forecast_metrics(code: str) -> Dict:
    """Project final outcomes based on current burn rate and sales pace."""
    report = calculate_report(code)
    metrics = calculate_inventory_metrics(code)
    ex = get_exhibition(code)
    
    # Calculate daily burn rate
    with connect() as conn:
        first_exp = conn.execute("SELECT MIN(created_at) FROM confirmed_expenses WHERE exhibition_code = ?", (code,)).fetchone()[0]
        if first_exp:
            start_dt = datetime.fromisoformat(first_exp.replace("Z", ""))
            days_active = max((datetime.utcnow() - start_dt).days, 1)
            total_exp = report["totals"]["direct_costs"] + report["totals"]["operating_expenses"] + report["totals"]["allocated_overhead"]
            daily_burn = total_exp / days_active
        else:
            daily_burn = 0.0
            
    # Simple projections
    # If we maintain the current sell-through rate, what is the projected revenue?
    potential_total_rev = report["totals"]["gallery_revenue"] + (metrics["unsold_asking_value_thb"] * (metrics["sell_through_rate_pct"] / 100))
    
    return {
        "daily_burn_rate": daily_burn,
        "projected_total_revenue": potential_total_rev,
        "projected_net_profit": potential_total_rev - (report["totals"]["direct_costs"] + report["totals"]["operating_expenses"] + report["totals"]["allocated_overhead"]),
        "days_active": days_active if first_exp else 0
    }

def calculate_inventory_metrics(exhibition_code: str) -> Dict:
    artworks = list_artworks(exhibition_code)
    sales = list_sales(exhibition_code)
    total_artworks = len(artworks)
    sold_artworks = len([row for row in artworks if row["status"] == "sold"])
    available_artworks = total_artworks - sold_artworks
    total_asking_value = sum(float(row["asking_price_thb"] or 0) for row in artworks)
    unsold_asking_value = sum(float(row["asking_price_thb"] or 0) for row in artworks if row["status"] != "sold")
    gross_sales = sum(float(row["actual_price_thb"] or 0) for row in sales)
    cash_collected = sum(float((row["amount_collected_thb"] if "amount_collected_thb" in row.keys() else row["actual_price_thb"]) or 0) for row in sales)
    receivables = sum(float((row["balance_due_thb"] if "balance_due_thb" in row.keys() else 0) or 0) for row in sales)
    average_sale_price = gross_sales / len(sales) if sales else 0.0
    sell_through = sold_artworks / total_artworks * 100 if total_artworks else 0.0
    return {
        "total_artworks": total_artworks,
        "sold_artworks": sold_artworks,
        "available_artworks": available_artworks,
        "total_asking_value_thb": total_asking_value,
        "unsold_asking_value_thb": unsold_asking_value,
        "gross_sales_thb": gross_sales,
        "cash_collected_thb": cash_collected,
        "receivables_thb": receivables,
        "average_sale_price_thb": average_sale_price,
        "sell_through_rate_pct": sell_through,
        "pending_receipts": count_pending_expenses(exhibition_code),
    }


def set_expense_budget(exhibition_code: str, account_head: str, budget_thb: float, notes: str | None = None) -> sqlite3.Row:
    exhibition_code = normalize_code(exhibition_code)
    if not get_exhibition(exhibition_code):
        raise ValueError(f"Exhibition not found: {exhibition_code}")
    account = get_account_head(account_head)
    amount = float(budget_thb)
    if amount < 0:
        raise ValueError("Budget amount cannot be negative.")
    now = _utc_now()
    with connect() as conn:
        conn.execute(
            """
            INSERT INTO expense_budgets (exhibition_code, account_head, section, budget_thb, notes, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(exhibition_code, account_head)
            DO UPDATE SET budget_thb = excluded.budget_thb, notes = excluded.notes, section = excluded.section, updated_at = excluded.updated_at
            """,
            (exhibition_code, account["name"], account["section"], amount, notes, now, now),
        )
        _insert_audit(conn, "set_expense_budget", exhibition_code, f"Set budget {account['name']} to {money(amount)}")
        return conn.execute(
            "SELECT * FROM expense_budgets WHERE UPPER(exhibition_code) = UPPER(?) AND account_head = ?",
            (exhibition_code, account["name"]),
        ).fetchone()


def list_expense_budgets(exhibition_code: str) -> List[sqlite3.Row]:
    with connect() as conn:
        return conn.execute(
            """
            SELECT * FROM expense_budgets
            WHERE UPPER(exhibition_code) = UPPER(?)
            ORDER BY account_head
            """,
            (exhibition_code,),
        ).fetchall()


def calculate_budget_report(exhibition_code: str) -> List[Dict]:
    confirmed = list_confirmed_expenses(exhibition_code) if get_exhibition(exhibition_code) else []
    budgets = {row["account_head"]: row for row in list_expense_budgets(exhibition_code)} if get_exhibition(exhibition_code) else {}
    names = sorted(set(account_head_names()) | set(row["account_head"] for row in confirmed) | set(budgets.keys()))
    rows: List[Dict] = []
    for name in names:
        account = ACCOUNT_HEAD_BY_NAME.get(name.lower()) or {"section": budgets.get(name, {}).get("section", "operating_expense") if budgets.get(name) else "operating_expense"}
        actual = sum(float(row["amount_thb"] or 0) for row in confirmed if row["account_head"] == name)
        budget = float(budgets[name]["budget_thb"] or 0) if name in budgets else 0.0
        variance = budget - actual
        utilization = actual / budget * 100 if budget else 0.0
        rows.append(
            {
                "account_head": name,
                "section": account["section"],
                "budget_thb": budget,
                "actual_thb": actual,
                "variance_thb": variance,
                "utilization_pct": utilization,
            }
        )
    return rows


def format_budget_report_markdown(exhibition_code: str) -> str:
    exhibition = get_exhibition(exhibition_code)
    if not exhibition:
        raise ValueError(f"Exhibition not found: {exhibition_code}")
    rows = calculate_budget_report(exhibition_code)
    budgeted = [row for row in rows if row["budget_thb"] or row["actual_thb"]]
    if not budgeted:
        return f"No budgets or confirmed expenses found for `{normalize_code(exhibition_code)}`. Set one with /budget {normalize_code(exhibition_code)} <account> <amount>."
    lines = [f"*Budget vs Actual — {exhibition['name']}*", ""]
    for row in budgeted:
        marker = "OVER" if row["budget_thb"] > 0 and row["variance_thb"] < 0 else "OK"
        lines.append(
            f"• {row['account_head']}: Actual {compact_money(row['actual_thb'])} / Budget {compact_money(row['budget_thb'])} / Variance {compact_money(row['variance_thb'])} ({marker})"
        )
    return "\n".join(lines)


def format_inventory_dashboard_markdown(exhibition_code: str) -> str:
    exhibition = get_exhibition(exhibition_code)
    if not exhibition:
        raise ValueError(f"Exhibition not found: {exhibition_code}")
    metrics = calculate_inventory_metrics(exhibition_code)
    lines = [
        f"*Inventory & Cash Dashboard — {exhibition['name']}*",
        "",
        f"Registered artworks: *{metrics['total_artworks']}*",
        f"Sold artworks: *{metrics['sold_artworks']}*",
        f"Available artworks: *{metrics['available_artworks']}*",
        f"Sell-through rate: *{metrics['sell_through_rate_pct']:.1f}%*",
        f"Total asking value: *{money(metrics['total_asking_value_thb'])}*",
        f"Unsold asking value: *{money(metrics['unsold_asking_value_thb'])}*",
        f"Gross sale value: *{money(metrics['gross_sales_thb'])}*",
        f"Average sale price: *{money(metrics['average_sale_price_thb'])}*",
        f"Cash collected: *{money(metrics['cash_collected_thb'])}*",
        f"Receivables outstanding: *{money(metrics['receivables_thb'])}*",
        f"Pending receipt approvals: *{metrics['pending_receipts']}*",
    ]
    return "\n".join(lines)


def format_readiness_markdown(exhibition_code: str) -> str:
    exhibition = get_exhibition(exhibition_code)
    if not exhibition:
        raise ValueError(f"Exhibition not found: {exhibition_code}")
    checks = data_quality_checks(exhibition_code)
    blocking = [c for c in checks if "No blocking issues" not in c]
    status = "READY FOR REVIEW" if not blocking else "NEEDS ATTENTION"
    lines = [
        f"*Final Review Readiness — {exhibition['name']}*",
        f"Status: *{status}*",
        "",
    ]
    lines.extend(f"• {check}" for check in checks)
    lines.append("")
    lines.append("Use the guided menu to clear pending receipts, set missing splits, correct budgets, or export the final workbook.")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------


def export_report_xlsx(code: str, output_dir: str = "./exports") -> str:
    """Export a full, formula-driven exhibition finance workbook.

    Convention (matches top-tier accounting-firm templates):
      - Arial = labels, section headers, and totals (bold where noted).
      - Calibri = calculated / linked cells (formula output).
      - Arial + blue font = a hardcoded manual input (e.g. a sale price or budget figure).
      - Yellow fill = a value that is missing, zero, or needs review.
      - Red bold (conditional format) = a variance that is over budget.
      - Every cross-sheet number is a real formula (SUMIF/SUM/COUNTA), not a
        Python-computed literal, so the workbook recalculates if edited.
    """
    from openpyxl.formatting.rule import CellIsRule

    report = calculate_report(code)
    ex = report["exhibition"]
    sales = list_sales(code)
    confirmed = list_confirmed_expenses(code)
    budgets = {row["account_head"]: row for row in list_expense_budgets(code)}
    artworks_all = list_artworks(code)
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    file_path = Path(output_dir) / f"{ex['code']}_pnl_report.xlsx"

    # ---- style kit -------------------------------------------------------
    MONEY_FMT = '#,##0.00;\\(#,##0.00\\);\\-'
    F_TITLE = Font(name="Arial", size=14, bold=True)
    F_SUBTITLE = Font(name="Arial", size=12, bold=True)
    F_LABEL = Font(name="Arial", size=11, bold=True)
    F_DATA = Font(name="Calibri", size=11, bold=False)
    F_TOTAL = Font(name="Arial", size=11, bold=True)
    F_TOTAL_BIG = Font(name="Arial", size=12, bold=True)
    F_INPUT = Font(name="Arial", size=11, bold=False, color="0000FF")
    F_NOTE = Font(name="Arial", size=9, bold=False)
    F_NOTE_BLUE = Font(name="Arial", size=9, bold=False, color="0000FF")
    F_NOTE_RED = Font(name="Arial", size=9, bold=False, color="C00000")
    FILL_YELLOW = PatternFill(fill_type="solid", fgColor="FFFF00")
    FILL_YELLOW_LIGHT = PatternFill(fill_type="solid", fgColor="FFFF99")
    AL_WRAP = Alignment(vertical="top", wrap_text=True)
    AL_WRAP_RIGHT = Alignment(vertical="top", wrap_text=True, horizontal="right")
    RED_BOLD_RULE = lambda: CellIsRule(operator="lessThan", formula=["0"], font=Font(bold=True, color="C00000"))

    def money(ws, coord, font=F_DATA):
        c = ws[coord]
        c.number_format = MONEY_FMT
        c.font = font
        c.alignment = AL_WRAP_RIGHT

    def label(ws, coord, text, font=F_DATA, align=None):
        c = ws[coord]
        c.value = text
        c.font = font
        c.alignment = align or AL_WRAP

    def header_row(ws, row, headers, start_col=1):
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=start_col + i, value=h)
            c.font = F_LABEL
            c.alignment = AL_WRAP

    def meta_block(ws, r):
        label(ws, f"A{r}", "Exhibition", F_LABEL); label(ws, f"B{r}", ex["name"]); label(ws, f"C{r}", "Code", F_LABEL); label(ws, f"D{r}", ex["code"])
        label(ws, f"A{r+1}", "Location", F_LABEL); label(ws, f"B{r+1}", ex["location"] or "\u2014"); label(ws, f"C{r+1}", "Currency", F_LABEL); label(ws, f"D{r+1}", "THB only")
        period = f"{ex['start_date'] or '?'} \u2192 {ex['end_date'] or 'ongoing'}"
        label(ws, f"A{r+2}", "Period", F_LABEL); label(ws, f"B{r+2}", period); label(ws, f"C{r+2}", "Generated", F_LABEL); label(ws, f"D{r+2}", datetime.now().strftime("%Y-%m-%d %H:%M"))
        return r + 4

    heads = account_head_names()
    wb = Workbook()

    # =======================================================================
    # 1) SALES LEDGER  (raw sale transactions — source of truth for revenue)
    # =======================================================================
    sl = wb.active
    sl.title = "Sales Ledger"
    label(sl, "A1", f"Sales Ledger \u2014 {ex['name']}", F_TITLE)
    label(sl, "A2", "One row per recorded sale. The P&L and Executive Summary pull from this sheet by formula.", F_NOTE)
    SL_HEAD_ROW = 4
    header_row(sl, SL_HEAD_ROW, ["Date", "Artwork", "Artist", "Buyer", "Sale Price (THB)", "Gallery Share (THB)",
                                  "Artist Payable (THB)", "Collaborator Share (THB)", "Collected (THB)", "Balance Due (THB)", "Status"])
    SL_DATA_START = SL_HEAD_ROW + 1
    r = SL_DATA_START
    for s in sales:
        label(sl, f"A{r}", s["sale_date"], F_DATA)
        label(sl, f"B{r}", s["title"], F_DATA)
        label(sl, f"C{r}", s["artist"], F_DATA)
        label(sl, f"D{r}", s["buyer_name"] or "\u2014", F_DATA)
        sl.cell(row=r, column=5, value=float(s["actual_price_thb"] or 0)); money(sl, f"E{r}", F_INPUT)
        sl.cell(row=r, column=6, value=float(s["gallery_share_thb"] or 0)); money(sl, f"F{r}", F_INPUT)
        sl.cell(row=r, column=7, value=float(s["artist_payable_thb"] or 0)); money(sl, f"G{r}", F_INPUT)
        sl.cell(row=r, column=8, value=float(s["collaborator_share_thb"] or 0)); money(sl, f"H{r}", F_INPUT)
        sl.cell(row=r, column=9, value=float(s["amount_collected_thb"] or 0)); money(sl, f"I{r}", F_INPUT)
        sl.cell(row=r, column=10, value=float(s["balance_due_thb"] or 0)); money(sl, f"J{r}", F_INPUT)
        label(sl, f"K{r}", (s["payment_status"] or "collected").title(), F_DATA)
        r += 1
    if sales:
        SL_DATA_END = r - 1
        total_row_sl = r
    else:
        label(sl, f"A{SL_DATA_START}", "No sales recorded for this exhibition yet.", F_NOTE)
        SL_DATA_END = SL_DATA_START
        total_row_sl = SL_DATA_START + 1
    label(sl, f"D{total_row_sl}", "TOTAL", F_TOTAL)
    for col in "EFGHIJ":
        sl[f"{col}{total_row_sl}"] = f"=SUM({col}{SL_DATA_START}:{col}{SL_DATA_END})"
        money(sl, f"{col}{total_row_sl}", F_TOTAL)
    sl.freeze_panes = f"A{SL_DATA_START}"
    for col, w in zip("ABCDEFGHIJK", [12, 26, 18, 18, 14, 14, 14, 16, 14, 14, 12]):
        sl.column_dimensions[col].width = w

    # =======================================================================
    # 2) ACTUAL LEDGER  (raw expense transactions)
    # =======================================================================
    al = wb.create_sheet("Actual Ledger")
    label(al, "A1", f"Actual Expense Ledger \u2014 {ex['name']}", F_TITLE)
    label(al, "A2", "One row per confirmed expense. Budget vs Actual and the P&L pull from this sheet by formula.", F_NOTE)
    AL_HEAD_ROW = 4
    header_row(al, AL_HEAD_ROW, ["Date", "Recipient", "Account Head", "Category", "Description / Notes", "Amount (THB)"])
    AL_DATA_START = AL_HEAD_ROW + 1
    r = AL_DATA_START
    for c in confirmed:
        label(al, f"A{r}", (c["created_at"] or "")[:10], F_DATA)
        label(al, f"B{r}", c["recipient"] if ("recipient" in c.keys() and c["recipient"]) else "\u2014", F_DATA)
        label(al, f"C{r}", c["account_head"], F_DATA)
        label(al, f"D{r}", c["category"] if ("category" in c.keys() and c["category"]) else "\u2014", F_DATA)
        label(al, f"E{r}", c["description"] or "", F_DATA)
        al.cell(row=r, column=6, value=float(c["amount_thb"] or 0)); money(al, f"F{r}", F_INPUT)
        r += 1
    if confirmed:
        AL_DATA_END = r - 1
        total_row_al = r
    else:
        label(al, f"A{AL_DATA_START}", "No confirmed expenses for this exhibition yet.", F_NOTE)
        AL_DATA_END = AL_DATA_START
        total_row_al = AL_DATA_START + 1
    label(al, f"E{total_row_al}", "TOTAL", F_TOTAL)
    al[f"F{total_row_al}"] = f"=SUM(F{AL_DATA_START}:F{AL_DATA_END})"
    money(al, f"F{total_row_al}", F_TOTAL)
    al.freeze_panes = f"A{AL_DATA_START}"
    for col, w in zip("ABCDEF", [12, 22, 28, 24, 38, 16]):
        al.column_dimensions[col].width = w

    # =======================================================================
    # 3) BUDGET VS ACTUAL
    # =======================================================================
    bv = wb.create_sheet("Budget vs Actual")
    label(bv, "A1", f"Budget vs Actual \u2014 {ex['name']}", F_TITLE)
    label(bv, "A2", "Blue = budget you set manually. Black = pulled live from the Actual Ledger. Yellow = no budget set yet.", F_NOTE)
    BV_HEAD_ROW = 4
    header_row(bv, BV_HEAD_ROW, ["Account Head", "Budget (THB)", "Actual (THB)", "Variance (THB)"])
    BV_DATA_START = BV_HEAD_ROW + 1
    for i, head in enumerate(heads):
        row = BV_DATA_START + i
        label(bv, f"A{row}", head, F_DATA)
        budget_val = float(budgets[head]["budget_thb"]) if head in budgets else 0.0
        bv.cell(row=row, column=2, value=budget_val)
        money(bv, f"B{row}", F_INPUT)
        if head not in budgets or budget_val == 0:
            bv[f"B{row}"].fill = FILL_YELLOW
        bv[f"C{row}"] = f"=SUMIF('Actual Ledger'!$C${AL_DATA_START}:$C${AL_DATA_END},A{row},'Actual Ledger'!$F${AL_DATA_START}:$F${AL_DATA_END})"
        money(bv, f"C{row}")
        bv[f"D{row}"] = f"=B{row}-C{row}"
        money(bv, f"D{row}")
    BV_DATA_END = BV_DATA_START + len(heads) - 1
    BV_TOTAL_ROW = BV_DATA_END + 1
    label(bv, f"A{BV_TOTAL_ROW}", "GRAND TOTAL", F_TOTAL)
    for col in "BCD":
        bv[f"{col}{BV_TOTAL_ROW}"] = f"=SUM({col}{BV_DATA_START}:{col}{BV_DATA_END})"
        money(bv, f"{col}{BV_TOTAL_ROW}", F_TOTAL)
    bv.conditional_formatting.add(f"D{BV_DATA_START}:D{BV_TOTAL_ROW}", RED_BOLD_RULE())

    categories = sorted(set((c["category"] for c in confirmed if "category" in c.keys() and c["category"])))
    cb_row = BV_TOTAL_ROW + 2
    label(bv, f"A{cb_row}", "Spend by Category (informational)", F_SUBTITLE)
    cb_head_row = cb_row + 1
    header_row(bv, cb_head_row, ["Category", "Actual (THB)"])
    cb_data_start = cb_head_row + 1
    if categories:
        for i, cat in enumerate(categories):
            row = cb_data_start + i
            label(bv, f"A{row}", cat, F_DATA)
            bv[f"B{row}"] = f"=SUMIF('Actual Ledger'!$D${AL_DATA_START}:$D${AL_DATA_END},A{row},'Actual Ledger'!$F${AL_DATA_START}:$F${AL_DATA_END})"
            money(bv, f"B{row}")
        cb_data_end = cb_data_start + len(categories) - 1
    else:
        label(bv, f"A{cb_data_start}", "No categories tagged on expenses yet.", F_NOTE)
        cb_data_end = cb_data_start

    legend_row = cb_data_end + 2
    label(bv, f"A{legend_row}", "Legend:", F_LABEL)
    label(bv, f"A{legend_row+1}", "Blue text = manually set budget input.", F_NOTE_BLUE)
    label(bv, f"A{legend_row+2}", "Yellow highlight = no budget set for this account head yet \u2014 shows \u09800 pending your input.", F_NOTE)
    bv[f"A{legend_row+2}"].fill = FILL_YELLOW_LIGHT
    label(bv, f"A{legend_row+3}", "Red variance = actual exceeded budget (over budget).", F_NOTE_RED)
    bv.freeze_panes = f"A{BV_DATA_START}"
    for col, w in zip("ABCD", [32, 16, 16, 16]):
        bv.column_dimensions[col].width = w

    # =======================================================================
    # 4) P&L
    # =======================================================================
    pl = wb.create_sheet("P&L")
    label(pl, "A1", "THE SEA ART GALLERY", F_TITLE)
    label(pl, "A2", "Expense Statement & Profit / Loss (Exhibition Production)", F_SUBTITLE)
    pr = meta_block(pl, 4) + 1
    label(pl, f"A{pr}", "A. Revenue", F_SUBTITLE); pr += 1
    label(pl, f"A{pr}", "Gross Artwork Sales", F_DATA)
    pl[f"C{pr}"] = f"=SUM('Sales Ledger'!E{SL_DATA_START}:E{SL_DATA_END})"; money(pl, f"C{pr}"); pr += 1
    label(pl, f"A{pr}", "Gallery Revenue (Commission)", F_LABEL)
    pl[f"C{pr}"] = f"=SUM('Sales Ledger'!F{SL_DATA_START}:F{SL_DATA_END})"; money(pl, f"C{pr}", F_TOTAL)
    total_revenue_row = pr; pr += 2

    label(pl, f"A{pr}", "B. Direct Production & Artist Costs", F_SUBTITLE); pr += 1
    direct_rows = []
    for head in heads:
        if ACCOUNT_HEAD_BY_NAME[head.lower()]["section"] != "direct_cost":
            continue
        label(pl, f"A{pr}", head, F_DATA)
        pl[f"C{pr}"] = f"=SUMIF('Budget vs Actual'!$A${BV_DATA_START}:$A${BV_DATA_END},A{pr},'Budget vs Actual'!$C${BV_DATA_START}:$C${BV_DATA_END})"
        money(pl, f"C{pr}")
        direct_rows.append(pr); pr += 1
    label(pl, f"A{pr}", "Artist Payable (from sold artwork)", F_DATA)
    pl[f"C{pr}"] = f"=SUM('Sales Ledger'!G{SL_DATA_START}:G{SL_DATA_END})"; money(pl, f"C{pr}")
    direct_rows.append(pr); pr += 1
    label(pl, f"A{pr}", "Collaborator / Collector Share", F_DATA)
    pl[f"C{pr}"] = f"=SUM('Sales Ledger'!H{SL_DATA_START}:H{SL_DATA_END})"; money(pl, f"C{pr}")
    direct_rows.append(pr); pr += 1
    label(pl, f"A{pr}", "Total Direct Costs", F_LABEL)
    pl[f"C{pr}"] = "=" + "+".join(f"C{x}" for x in direct_rows); money(pl, f"C{pr}", F_TOTAL)
    total_direct_row = pr; pr += 2

    label(pl, f"A{pr}", "C. Operating Expenses", F_SUBTITLE); pr += 1
    opex_rows = []
    for head in heads:
        if ACCOUNT_HEAD_BY_NAME[head.lower()]["section"] != "operating_expense":
            continue
        label(pl, f"A{pr}", head, F_DATA)
        pl[f"C{pr}"] = f"=SUMIF('Budget vs Actual'!$A${BV_DATA_START}:$A${BV_DATA_END},A{pr},'Budget vs Actual'!$C${BV_DATA_START}:$C${BV_DATA_END})"
        money(pl, f"C{pr}")
        opex_rows.append(pr); pr += 1
    label(pl, f"A{pr}", "Total Operating Expenses", F_LABEL)
    pl[f"C{pr}"] = ("=" + "+".join(f"C{x}" for x in opex_rows)) if opex_rows else "=0"
    money(pl, f"C{pr}", F_TOTAL)
    total_opex_row = pr; pr += 2

    label(pl, f"A{pr}", "NET PROFIT / (LOSS)", F_TOTAL_BIG)
    pl[f"C{pr}"] = f"=C{total_revenue_row}-C{total_direct_row}-C{total_opex_row}"
    money(pl, f"C{pr}", F_TOTAL_BIG)
    pr += 2
    label(pl, f"A{pr}", "Direct and Operating costs above reconcile to the Actual Ledger via Budget vs Actual, and revenue reconciles to the Sales Ledger.", F_NOTE)
    pl.freeze_panes = "A9"
    for col, w in zip("ABCD", [40, 24, 18, 14]):
        pl.column_dimensions[col].width = w

    # =======================================================================
    # 5) PAINTINGS LIST
    # =======================================================================
    pw = wb.create_sheet("Paintings List")
    label(pw, "A1", f"Paintings List \u2014 {ex['name']}", F_TITLE)
    label(pw, "A2", "Full artwork inventory for this exhibition, sold or available.", F_NOTE)
    PW_HEAD_ROW = 4
    header_row(pw, PW_HEAD_ROW, ["No.", "Artist", "Title", "Medium", "Dimensions", "Year", "Asking Price (THB)", "Status"])
    PW_DATA_START = PW_HEAD_ROW + 1
    r = PW_DATA_START
    for idx, a in enumerate(artworks_all, start=1):
        label(pw, f"A{r}", idx, F_DATA)
        label(pw, f"B{r}", a["artist"], F_DATA)
        label(pw, f"C{r}", a["title"], F_DATA)
        label(pw, f"D{r}", a["medium"] if ("medium" in a.keys() and a["medium"]) else "\u2014", F_DATA)
        label(pw, f"E{r}", a["dimensions"] if ("dimensions" in a.keys() and a["dimensions"]) else "\u2014", F_DATA)
        label(pw, f"F{r}", a["year_created"] if ("year_created" in a.keys() and a["year_created"]) else "\u2014", F_DATA)
        pw.cell(row=r, column=7, value=float(a["asking_price_thb"] or 0)); money(pw, f"G{r}", F_INPUT)
        label(pw, f"H{r}", a["status"], F_DATA)
        r += 1
    if artworks_all:
        PW_DATA_END = r - 1
    else:
        label(pw, f"A{PW_DATA_START}", "No artworks registered for this exhibition yet.", F_NOTE)
        PW_DATA_END = PW_DATA_START
        r = PW_DATA_START + 1
    label(pw, f"F{r}", "Total Asking Value (THB)", F_TOTAL)
    pw[f"G{r}"] = f"=SUM(G{PW_DATA_START}:G{PW_DATA_END})"; money(pw, f"G{r}", F_TOTAL)
    label(pw, f"F{r+1}", "Number of Paintings", F_TOTAL)
    pw[f"G{r+1}"] = f"=COUNTA(B{PW_DATA_START}:B{PW_DATA_END})"
    pw[f"G{r+1}"].font = F_TOTAL
    pw.freeze_panes = f"A{PW_DATA_START}"
    for col, w in zip("ABCDEFGH", [6, 22, 32, 24, 16, 8, 18, 12]):
        pw.column_dimensions[col].width = w

    # =======================================================================
    # 6) EXECUTIVE SUMMARY  (built last so it can point at real row numbers)
    # =======================================================================
    es = wb.create_sheet("Executive Summary")
    wb.move_sheet("Executive Summary", offset=-(len(wb.sheetnames) - 1))
    label(es, "A1", "THE SEA ART GALLERY", F_TITLE)
    label(es, "A2", "Elite Exhibition Finance Summary", F_SUBTITLE)
    r = meta_block(es, 4) + 1

    label(es, f"A{r}", "Budget vs Actual by Account Head", F_SUBTITLE); r += 1
    header_row(es, r, ["Account Head", "Budget (THB)", "Actual (THB)", "Variance (THB)"]); r += 1
    es_data_start = r
    for head in heads:
        label(es, f"A{r}", head, F_DATA)
        es[f"B{r}"] = f"=SUMIF('Budget vs Actual'!$A${BV_DATA_START}:$A${BV_DATA_END},A{r},'Budget vs Actual'!$B${BV_DATA_START}:$B${BV_DATA_END})"
        money(es, f"B{r}")
        es[f"C{r}"] = f"=SUMIF('Budget vs Actual'!$A${BV_DATA_START}:$A${BV_DATA_END},A{r},'Budget vs Actual'!$C${BV_DATA_START}:$C${BV_DATA_END})"
        money(es, f"C{r}")
        es[f"D{r}"] = f"=B{r}-C{r}"
        money(es, f"D{r}")
        r += 1
    es_data_end = r - 1
    label(es, f"A{r}", "TOTAL", F_TOTAL)
    for col in "BCD":
        es[f"{col}{r}"] = f"=SUM({col}{es_data_start}:{col}{es_data_end})"
        money(es, f"{col}{r}", F_TOTAL)
    es.conditional_formatting.add(f"D{es_data_start}:D{r}", RED_BOLD_RULE())
    r += 2

    label(es, f"A{r}", "Revenue & Profitability", F_SUBTITLE); r += 1
    label(es, f"A{r}", "Gross Artwork Sales", F_DATA)
    es[f"B{r}"] = f"=SUM('Sales Ledger'!E{SL_DATA_START}:E{SL_DATA_END})"; money(es, f"B{r}"); r += 1
    label(es, f"A{r}", "Gallery Revenue (Commission)", F_DATA)
    es[f"B{r}"] = f"='P&L'!C{total_revenue_row}"; money(es, f"B{r}"); r += 1
    label(es, f"A{r}", "Total Direct + Operating Costs", F_DATA)
    es[f"B{r}"] = f"='P&L'!C{total_direct_row}+'P&L'!C{total_opex_row}"; money(es, f"B{r}"); r += 1
    label(es, f"A{r}", "Net Profit / (Loss)", F_TOTAL_BIG)
    es[f"B{r}"] = f"='P&L'!C{total_revenue_row}-'P&L'!C{total_direct_row}-'P&L'!C{total_opex_row}"
    money(es, f"B{r}", F_TOTAL_BIG); r += 2

    label(es, f"A{r}", "Inventory", F_SUBTITLE); r += 1
    label(es, f"A{r}", "Number of Paintings", F_DATA)
    es[f"B{r}"] = f"=COUNTA('Paintings List'!B{PW_DATA_START}:B{PW_DATA_END})"; r += 1
    label(es, f"A{r}", "Total Asking Value (THB)", F_DATA)
    es[f"B{r}"] = f"=SUM('Paintings List'!G{PW_DATA_START}:G{PW_DATA_END})"; money(es, f"B{r}"); r += 2

    label(es, f"A{r}", "Every figure above is a live formula. Edit the Sales Ledger, Actual Ledger, or Budget vs Actual sheets and this summary recalculates.", F_NOTE)
    es.freeze_panes = f"A{es_data_start}"
    for col, w in zip("ABCD", [36, 20, 20, 20]):
        es.column_dimensions[col].width = w

    # =======================================================================
    # 7) ARTIST PAYABLES
    # =======================================================================
    ap = wb.create_sheet("Artist Payables")
    label(ap, "A1", f"Artist Payables \u2014 {ex['name']}", F_TITLE)
    header_row(ap, 3, ["Artist", "Invoice Ref", "Gross Sale (THB)", "Gallery Commission (THB)", "Artist Payable (THB)", "Paid (THB)", "Outstanding (THB)", "Status"])
    r = 4
    for row in report["payables"]:
        label(ap, f"A{r}", row["artist"], F_DATA)
        label(ap, f"B{r}", row.get("invoice_ref") or "", F_DATA)
        ap.cell(row=r, column=3, value=float(row["gross_sale_thb"])); money(ap, f"C{r}")
        ap.cell(row=r, column=4, value=float(row["gallery_commission_thb"])); money(ap, f"D{r}")
        ap.cell(row=r, column=5, value=float(row["artist_payable_thb"])); money(ap, f"E{r}")
        ap.cell(row=r, column=6, value=float(row["paid_thb"])); money(ap, f"F{r}")
        ap.cell(row=r, column=7, value=float(row["outstanding_thb"])); money(ap, f"G{r}")
        label(ap, f"H{r}", row["status"], F_DATA)
        r += 1
    if not report["payables"]:
        label(ap, "A4", "No artist payables recorded yet.", F_NOTE)
    ap.freeze_panes = "A4"
    for col, w in zip("ABCDEFGH", [22, 24, 16, 18, 16, 14, 16, 12]):
        ap.column_dimensions[col].width = w

    # =======================================================================
    # 8) ARTIST ROI MATRIX
    # =======================================================================
    roi_ws = wb.create_sheet("Artist ROI Matrix")
    label(roi_ws, "A1", f"Artist ROI Matrix \u2014 {ex['name']}", F_TITLE)
    header_row(roi_ws, 3, ["Artist", "Total Works", "Sold", "Gallery Share (THB)", "Direct Costs (THB)", "Net Contribution (THB)", "Sell-through %"])
    r = 4
    artist_roi_rows = calculate_artist_roi(code)
    for rr in artist_roi_rows:
        label(roi_ws, f"A{r}", rr["artist"], F_DATA)
        label(roi_ws, f"B{r}", rr["total_artworks"], F_DATA)
        label(roi_ws, f"C{r}", rr["sold_artworks"], F_DATA)
        roi_ws.cell(row=r, column=4, value=float(rr["gallery_share"])); money(roi_ws, f"D{r}")
        roi_ws.cell(row=r, column=5, value=float(rr["direct_costs"])); money(roi_ws, f"E{r}")
        roi_ws.cell(row=r, column=6, value=float(rr["net_contribution"])); money(roi_ws, f"F{r}")
        roi_ws.cell(row=r, column=7, value=round(float(rr["sell_through_rate"]), 1))
        roi_ws[f"G{r}"].number_format = '0.0"%"'
        roi_ws[f"G{r}"].font = F_DATA
        r += 1
    if not artist_roi_rows:
        label(roi_ws, "A4", "No artists registered yet.", F_NOTE)
    roi_ws.freeze_panes = "A4"
    for col, w in zip("ABCDEFG", [22, 12, 10, 16, 16, 18, 14]):
        roi_ws.column_dimensions[col].width = w

    # =======================================================================
    # 9) RECEIVABLES AGING
    # =======================================================================
    rec_ws = wb.create_sheet("Receivables Aging")
    label(rec_ws, "A1", f"Receivables Aging \u2014 {ex['name']}", F_TITLE)
    header_row(rec_ws, 3, ["Artwork", "Artist", "Buyer", "Sale Price (THB)", "Collected (THB)", "Balance Due (THB)", "Status"])
    r = 4
    for s in sales:
        if float(s["balance_due_thb"] or 0) <= 0.009:
            continue
        label(rec_ws, f"A{r}", s["title"], F_DATA)
        label(rec_ws, f"B{r}", s["artist"], F_DATA)
        label(rec_ws, f"C{r}", s["buyer_name"] or "\u2014", F_DATA)
        rec_ws.cell(row=r, column=4, value=float(s["actual_price_thb"])); money(rec_ws, f"D{r}")
        rec_ws.cell(row=r, column=5, value=float(s["amount_collected_thb"] or 0)); money(rec_ws, f"E{r}")
        rec_ws.cell(row=r, column=6, value=float(s["balance_due_thb"] or 0)); money(rec_ws, f"F{r}")
        label(rec_ws, f"G{r}", (s["payment_status"] or "").title(), F_DATA)
        r += 1
    if r == 4:
        label(rec_ws, "A4", "No outstanding receivables \u2014 every recorded sale is fully collected.", F_NOTE)
    rec_ws.freeze_panes = "A4"
    for col, w in zip("ABCDEFG", [26, 20, 20, 16, 16, 16, 14]):
        rec_ws.column_dimensions[col].width = w

    # =======================================================================
    # 10) AUDIT LOG
    # =======================================================================
    audit = wb.create_sheet("Audit Log")
    label(audit, "A1", f"Audit Log \u2014 {ex['name']}", F_TITLE)
    header_row(audit, 3, ["Timestamp", "Action", "Exhibition", "Details"])
    with connect() as conn:
        audit_rows = conn.execute(
            """
            SELECT timestamp, action, exhibition_code, details
            FROM audit_log
            WHERE exhibition_code IS NULL OR UPPER(exhibition_code) = UPPER(?)
            ORDER BY id DESC LIMIT 250
            """,
            (code,),
        ).fetchall()
    r = 4
    for row in audit_rows:
        label(audit, f"A{r}", row["timestamp"], F_DATA)
        label(audit, f"B{r}", row["action"], F_DATA)
        label(audit, f"C{r}", row["exhibition_code"] or "", F_DATA)
        label(audit, f"D{r}", row["details"] or "", F_DATA)
        r += 1
    if not audit_rows:
        label(audit, "A4", "No audit entries yet.", F_NOTE)
    audit.freeze_panes = "A4"
    for col, w in zip("ABCD", [20, 24, 16, 60]):
        audit.column_dimensions[col].width = w

    # ---- global finishing pass: thin border on every used cell ----------
    thin = Side(style="thin", color="000000")
    border_all = Border(top=thin, left=thin, right=thin, bottom=thin)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border_all
                if cell.alignment is None or cell.alignment.horizontal is None:
                    cell.alignment = AL_WRAP

    wb.save(file_path)
    log_action("export_xlsx", code, str(file_path))
    return str(file_path)



# ---------------------------------------------------------------------------
# Payment collection / artist settlement helpers
# ---------------------------------------------------------------------------


def collect_sale_payment(sale_id: int, amount_thb: float, notes: str | None = None) -> sqlite3.Row:
    """Record an additional cash collection against an existing sale's balance due."""
    amount = float(amount_thb)
    if amount <= 0:
        raise ValueError("Collection amount must be greater than zero.")
    with connect() as conn:
        sale = conn.execute("SELECT * FROM artwork_sales WHERE id = ?", (int(sale_id),)).fetchone()
        if not sale:
            raise ValueError(f"Sale not found: {sale_id}")
        current_collected = float(sale["amount_collected_thb"] or 0)
        actual_price = float(sale["actual_price_thb"] or 0)
        new_collected = round(current_collected + amount, 2)
        if new_collected > actual_price + 0.01:
            raise ValueError("Collection would exceed the actual sale price.")
        new_balance = round(actual_price - new_collected, 2)
        status = "collected" if new_balance <= 0.01 else "partial"
        merged_notes = ((sale["notes"] or "") + f" | +{money(amount)} collected {_utc_now()}").strip(" |")
        conn.execute(
            """
            UPDATE artwork_sales
            SET amount_collected_thb = ?, balance_due_thb = ?, payment_status = ?, notes = ?
            WHERE id = ?
            """,
            (new_collected, new_balance, status, merged_notes if notes is None else f"{merged_notes}; {notes}", int(sale_id)),
        )
        _insert_audit(conn, "collect_sale_payment", sale["exhibition_code"], f"Collected {money(amount)} against sale #{sale_id}; new balance {money(new_balance)}")
        return conn.execute("SELECT * FROM artwork_sales WHERE id = ?", (int(sale_id),)).fetchone()


def record_artist_payment(payable_id: int, amount_thb: float) -> sqlite3.Row:
    """Record a payment made to an artist against their outstanding payable."""
    amount = float(amount_thb)
    if amount <= 0:
        raise ValueError("Payment amount must be greater than zero.")
    with connect() as conn:
        row = conn.execute("SELECT * FROM artist_payables WHERE id = ?", (int(payable_id),)).fetchone()
        if not row:
            raise ValueError(f"Artist payable not found: {payable_id}")
        payable_total = float(row["artist_payable_thb"] or 0)
        current_paid = float(row["paid_thb"] or 0)
        new_paid = round(current_paid + amount, 2)
        if new_paid > payable_total + 0.01:
            raise ValueError("Payment would exceed the total artist payable.")
        new_outstanding = round(payable_total - new_paid, 2)
        status = "Paid" if new_outstanding <= 0.01 else "Pending"
        conn.execute(
            "UPDATE artist_payables SET paid_thb = ?, outstanding_thb = ?, status = ? WHERE id = ?",
            (new_paid, new_outstanding, status, int(payable_id)),
        )
        _insert_audit(conn, "record_artist_payment", row["exhibition_code"], f"Paid {money(amount)} to {row['artist']} (payable #{payable_id}); outstanding {money(new_outstanding)}")
        return conn.execute("SELECT * FROM artist_payables WHERE id = ?", (int(payable_id),)).fetchone()


def get_exhibition_closeout_status(code: str) -> Dict:
    with connect() as conn:
        ex = get_exhibition(code)
        if not ex:
            raise ValueError(f"Exhibition not found: {code}")

        artworks_total = conn.execute("SELECT COUNT(*) FROM artworks WHERE exhibition_code=?", (code,)).fetchone()[0]
        artworks_sold = conn.execute("SELECT COUNT(*) FROM artworks WHERE exhibition_code=? AND status='sold'", (code,)).fetchone()[0]
        artworks_avail = artworks_total - artworks_sold

        pending_count = conn.execute("SELECT COUNT(*) FROM pending_expenses WHERE exhibition_code=? AND status='pending'", (code,)).fetchone()[0]

        unpaid_artists = conn.execute("SELECT COUNT(*) FROM artist_payables WHERE exhibition_code=? AND outstanding_thb > 0.01", (code,)).fetchone()[0]
        unpaid_artists_list = conn.execute("SELECT * FROM artist_payables WHERE exhibition_code=? AND outstanding_thb > 0.01", (code,)).fetchall()

        outstanding_receivables = conn.execute("SELECT SUM(balance_due_thb) FROM artwork_sales WHERE exhibition_code=?", (code,)).fetchone()[0] or 0.0

        return {
            "exhibition": dict(ex),
            "artworks_total": artworks_total,
            "artworks_sold": artworks_sold,
            "artworks_avail": artworks_avail,
            "pending_count": pending_count,
            "unpaid_artists_count": unpaid_artists,
            "unpaid_artists": [dict(r) for r in unpaid_artists_list],
            "outstanding_receivables": outstanding_receivables,
        }


def close_exhibition_in_db(code: str) -> None:
    with connect() as conn:
        conn.execute("UPDATE exhibitions SET status = 'completed', end_date = ? WHERE code = ?", (datetime.now().strftime("%Y-%m-%d"), code))
        _insert_audit(conn, "close_exhibition", code, "Exhibition closed out and status set to completed.")


def get_cash_flow_timeline(code: str) -> List[Dict]:
    timeline = []
    with connect() as conn:
        sales = conn.execute(
            """
            SELECT s.sale_date as date, a.title, s.actual_price_thb, s.amount_collected_thb
            FROM artwork_sales s
            LEFT JOIN artworks a ON a.id = s.artwork_id
            WHERE UPPER(s.exhibition_code) = UPPER(?)
            """, (code,)
        ).fetchall()
        for r in sales:
            timeline.append({
                "type": "sale",
                "date": r["date"],
                "description": f"Sale: {r['title']}",
                "amount": float(r["amount_collected_thb"] or 0),
                "total_value": float(r["actual_price_thb"] or 0),
            })

        expenses = conn.execute(
            """
            SELECT SUBSTR(created_at, 1, 10) as date, description, amount_thb, account_head
            FROM confirmed_expenses
            WHERE UPPER(exhibition_code) = UPPER(?)
            """, (code,)
        ).fetchall()
        for r in expenses:
            timeline.append({
                "type": "expense",
                "date": r["date"],
                "description": f"{r['account_head']} - {r['description']}",
                "amount": -float(r["amount_thb"] or 0),
            })

    timeline.sort(key=lambda x: x["date"])
    return timeline


# ===========================================================================
# User State Management (DB Persisted) — kept for Telegram-era compatibility
# ===========================================================================

import json

def get_user_state(chat_id: int) -> dict:
    with connect() as conn:
        row = conn.execute("SELECT * FROM user_states WHERE chat_id = ?", (chat_id,)).fetchone()
        if row:
            return dict(row)
        default_exh = resolve_default_exhibition()
        return {
            "chat_id": chat_id,
            "current_exhibition": default_exh,
            "active_flow": None,
            "flow_step": 0,
            "flow_data": "{}",
        }

def resolve_default_exhibition() -> str:
    try:
        with connect() as conn:
            row = conn.execute(
                "SELECT code FROM exhibitions WHERE status NOT IN ('prototype', 'completed') "
                "ORDER BY COALESCE(end_date, start_date, '9999') DESC, code LIMIT 1"
            ).fetchone()
            if row:
                return row[0]
            row = conn.execute("SELECT code FROM exhibitions ORDER BY rowid DESC LIMIT 1").fetchone()
            if row:
                return row[0]
    except Exception:
        pass
    raw_default = os.environ.get("DEFAULT_EXHIBITION", "")
    return re.sub(r"[^A-Z0-9_]", "", raw_default.split()[0].upper()) if raw_default else ""



def seed_hervalor_if_missing(path: str | None = None) -> None:
    """Ensure HERVALOR2026 exhibition data exists and is seeded automatically."""
    with connect(path) as conn:
        row = conn.execute("SELECT code FROM exhibitions WHERE UPPER(code) = 'HERVALOR2026'").fetchone()
        if row:
            # Update status and splits if existing
            conn.execute("UPDATE exhibitions SET status = 'completed' WHERE UPPER(code) = 'HERVALOR2026'")
            conn.execute("DELETE FROM commission_splits WHERE UPPER(exhibition_code) = 'HERVALOR2026'")
            conn.execute("INSERT INTO commission_splits (exhibition_code, party_type, party_name, percent) VALUES ('HERVALOR2026', 'gallery', 'Gallery', 15)")
            conn.execute("INSERT INTO commission_splits (exhibition_code, party_type, party_name, percent) VALUES ('HERVALOR2026', 'artist', 'Artist', 85)")
            return

        conn.execute("""
            INSERT OR REPLACE INTO exhibitions (code, name, location, start_date, end_date, status, currency)
            VALUES ('HERVALOR2026', 'Her Valor, Her Vision, Her Voice', 'Chiang Mai University Art Center', '2026-05-11', '2026-06-28', 'completed', 'THB')
        """)

        conn.execute("DELETE FROM commission_splits WHERE UPPER(exhibition_code) = 'HERVALOR2026'")
        conn.execute("INSERT INTO commission_splits (exhibition_code, party_type, party_name, percent) VALUES ('HERVALOR2026', 'gallery', 'Gallery', 15)")
        conn.execute("INSERT INTO commission_splits (exhibition_code, party_type, party_name, percent) VALUES ('HERVALOR2026', 'artist', 'Artist', 85)")

        artworks = [
            ('No.', 'Source: "List of Paintings" sheet, Artists & Artworks Tracking workbook', 80000, 'available'),
            ('Size', 'Year', 13000, 'available'),
            ('Asking Price (THB)', 'Year', 13000, 'available'),
            ('Boat Sutasinee', 'Year', 14000, 'available'),
            ('Not Finished', 'Year', 14000, 'available'),
            ('120 x 80 cm', 'Year', 14000, 'available'),
            ('Mixed Media', 'Year', 14000, 'available'),
            ('Chuu Wai', 'Year', 14000, 'available'),
            ('Sleep in peace', 'Year', 14000, 'available'),
            ('20 × 20 cm', 'Year', 13000, 'available'),
            ('HOME?', 'Year', 14000, 'available'),
            ('Woven Dilemma 1', 'Year', 14000, 'available'),
            ('Woven Dilemma 2', 'Year', 14000, 'available'),
            ('Woven Dilemma 3', 'Year', 14000, 'available'),
            ('Woven Dilemma 4', 'Year', 22000, 'available'),
            ('Woven Dilemma 6', 'Year', 14000, 'available'),
            ('WOVEN DILEMMA 7', 'Year', 14000, 'available'),
            ('Woven Dilemma 8', 'Year', 22000, 'available'),
            ('Woven Dilemma 9', 'Year', 22000, 'available'),
            ('Woven Dilemma 11', 'Woven Dilemma 10', 22400, 'available'),
            ('Woven Dilemma 14', 'Woven Dilemma 10', 22400, 'available'),
            ('ECHOES OF THE LOOM', 'Woven Dilemma 15', 180000, 'available'),
            ('Witches 2', 'Woven Dilemma 15', 80000, 'available'),
            ('Witches 1', 'Woven Dilemma 15', 80000, 'available'),
            ('Kan Nathiwutthikun', 'Woven Dilemma 15', 80000, 'available'),
            ('155.5 × 156.5 cm', 'Woven Dilemma 15', 80000, 'available'),
            ('Untitled #wm2-1', 'Acrylic on Canvas', 16000, 'available'),
            ('300 x 120 cm', 'Acrylic on Canvas', 16000, 'available'),
            ('Acrylic on Fabric', 'Acrylic on Canvas', 10000, 'available'),
            ('100 x 80 cm', 'Untitled#Ep1', 30000, 'available'),
            ('Khin Khin Aye', 'Untitled#Ep1', 30000, 'available'),
            ('Inside Out Beauty', 'Sleeping Ugly', 128000, 'available'),
            ('Kyu Kyu', '50 × 60 cm', 80000, 'available'),
            ('101 x 76 cm', '50 × 60 cm', 96000, 'available'),
            ('Ma Thi', 'Silent 2', 89000, 'available'),
            ('Spinning Yarn Under the Moonlight', 'Nann Nann', 67000, 'available'),
            ('Spinning Yarn Under the Moonlight', 'Nann Nann', 0, 'available'),
            ('Peace be upon you', 'Nitaya Ueareeworakul', 20000, 'available'),
            ('PAEN (PANDA)', 'Nitaya Ueareeworakul', 20000, 'available'),
            ('30 x 30 cm', 'Who When What', 35000, 'available'),
            ('80 x 100 cm', 'Between day and night 1', 48000, 'available'),
            ('Pattree Chimnok', 'Between day and night 2', 82500, 'available'),
            ('100 x 130 cm', 'Body & Earth "when Silent Speaks', 120000, 'available'),
            ('Phi Phi', 'Body & Earth "when Silent Speaks', 75000, 'available'),
            ('Sandar Khaing', '107 x 152 cm', 15000, 'available'),
            ('Inherited Silence', '107 x 152 cm', 15000, 'available'),
            ('Study on Guardian Angel No 1', 'Sudaporn Teja', 65432, 'available'),
            ('Study on Guardian Angel No 2', 'Golden Teak Sawdust and Acrylic on Canvas', 57600, 'available'),
        ]
        for title, artist, price, status in artworks:
            conn.execute("INSERT INTO artworks (exhibition_code, title, artist, asking_price_thb, status) VALUES ('HERVALOR2026', ?, ?, ?, ?)", (title, artist, price, status))

        expenses = [
            (9180, 'Venue Rental', 'Rental Fees', 'Advance Payment', 'CMU Art Center', '2026-05-11'),
            (21420, 'Venue Rental', 'Rental Fees', 'Second Payment', 'CMU Art Center', '2026-06-03'),
            (5000, 'Food & Beverage / Hospitality', 'Refreshment', 'Food and Beverages', 'Catering Service', '2026-06-09'),
            (5028, 'Food & Beverage / Hospitality', 'Refreshment', 'Closing Dinner', 'Samsen Restaurant', '2026-06-18'),
            (180, 'Venue Rental', 'Venue Services', 'Overtime charges', 'CMU Art Center', '2026-06-09'),
            (3078, 'Framing & Artwork Preparation', 'Artwork Handling', 'Artwork Transportation from Yangon', 'Air Cargo', '2026-06-16'),
            (178, 'Framing & Artwork Preparation', 'Artwork Handling', 'Artwork Transportation to Chiang Mai', 'BKK Post Office', '2026-06-17'),
            (1100, 'Framing & Artwork Preparation', 'Artwork Handling', 'Artwork Transportation to Art Center', 'Car Rental Service', '2026-06-08'),
            (1100, 'Framing & Artwork Preparation', 'Artwork Handling', 'Artwork Transportation to Art Center', 'Car Rental Service', '2026-06-19'),
            (1066, 'Framing & Artwork Preparation', 'Artwork Handling', 'Bubble Wraps and Tapes', 'Mr. DIY', '2026-06-20'),
            (1000, 'Installation & Production', 'Installation and Dismantling', 'Installing', 'Installing Team', '2026-06-08'),
            (750, 'Installation & Production', 'Installation and Dismantling', 'Dismantling', 'Installing Team', '2026-06-19'),
            (3000, 'Venue Rental', 'Utilities', 'Mobile Aircon Rental', 'Air-Con Service', '2026-06-09'),
            (3646.78, 'Travel & Accommodation', 'Artist/Curator Travel', '3 Seats of Bus to Chiang Mai', 'Bus', '2026-06-08'),
            (1593.6, 'Travel & Accommodation', 'Artist/Curator Travel', '2 Seats of Bus to Chiang Mai', 'Bus', '2026-06-20'),
            (10000, 'Travel & Accommodation', 'Artist/Curator Travel', 'Hotel Room Charges', 'B House Hotel', '2026-06-22'),
            (2941, 'Installation & Production', 'Signage and Labels', 'Printing of Artists\' Biography', 'Printing House', '2026-06-10'),
            (500, 'Installation & Production', 'Display Materials', 'Fabrics', 'Fabrics Vendor', '2026-06-08'),
            (3350, 'Framing & Artwork Preparation', 'Artwork Handling', 'Artwork Framing', 'Framing Service', '2026-06-25'),
            (74111, '', 'Miscellaneous (Needs Review)', 'Printed total on source PDF', '', '2026-05-11'),
        ]
        for amt, desc, head, cat, rec, dt in expenses:
            conn.execute("INSERT INTO expenses_confirmed (exhibition_code, amount_thb, description, account_head, category, recipient, created_at) VALUES ('HERVALOR2026', ?, ?, ?, ?, ?, ?)", (amt, desc, head, cat, rec, dt))

        budgets = [
            ('Venue Rental', 33780),
            ('Framing & Artwork Preparation', 9872),
            ('Installation & Production', 5191),
            ('Marketing & PR', 0),
            ('Staff & Helpers / Labor', 0),
            ('Travel & Accommodation', 15240),
            ('Food & Beverage / Hospitality', 0),
            ('Office & Admin Supplies', 0),
            ('Miscellaneous (Needs Review)', 0),
        ]
        for head, b_amt in budgets:
            conn.execute("INSERT INTO expense_budgets (exhibition_code, account_head, budget_thb) VALUES ('HERVALOR2026', ?, ?)", (head, b_amt))
